#!/usr/bin/env python3
"""
Azure Region BOM Checker
========================
Reads an Azure Bill of Materials from bom_template.xlsx (or a specified file),
then checks specified Azure regions to determine whether they support all listed services.

Special checks:
  • Azure Automation  – not available in every region; queried via resource provider
  • Premium SSD v2    – requires zone-level availability; queried via az vm list-skus

Usage:
  python check_azure_regions.py --regions eastus,mexicocentral
  python check_azure_regions.py --bom bom_template.xlsx --regions eastus,westeurope --output results.xlsx
  python check_azure_regions.py --regions eastus,centralus,westus2 --verbose
  python check_azure_regions.py --regions-file regions.txt
  python check_azure_regions.py --regions-file regions.txt --regions eastus,westus3
  python check_azure_regions.py --regions all        # check every Azure region (slow)

Regions file format (--regions-file):
  • One region per line (e.g. `eastus`)
  • Lines starting with `#` are treated as comments
  • Blank lines are ignored
  • Comma-separated tokens on a single line are also accepted
  • `--regions-file` and `--regions` may be combined (results are deduped)

Prerequisites:
  pip install openpyxl
  az login   (Azure CLI must be installed and authenticated)

Disclaimer:
  This tool was created with the assistance of GitHub Copilot as a personal project.
  It is not an official Microsoft tool and is not endorsed or supported by Microsoft.
  Review all code before running it in your environment.
"""

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# Ensure UTF-8 output on Windows so emoji status symbols display correctly
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except AttributeError:
        pass  # Python < 3.7 fallback

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed.  Run: pip install openpyxl")

# ── Cell fill colours ──────────────────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # all required services available
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # available but SSDv2 has <3 zones
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")   # one or more required services missing
FILL_GRAY   = PatternFill("solid", fgColor="D9D9D9")   # optional / not required
FILL_HDR    = PatternFill("solid", fgColor="1F4E79")
FILL_KEY    = PatternFill("solid", fgColor="FFD966")
FILL_BLUE   = PatternFill("solid", fgColor="BDD7EE")

FONT_HDR    = Font(bold=True, color="FFFFFF", size=10)
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(vertical="center", wrap_text=True)
thin   = Side(style="thin", color="9DC3E6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ── Azure CLI helpers ──────────────────────────────────────────────────────────
def az(args: list[str], verbose: bool = False) -> dict | list | None:
    """Run an Azure CLI command and return parsed JSON output.
    Uses shell=True so that az.cmd is resolved correctly on Windows.
    Arguments containing shell-special characters are wrapped in double quotes."""
    _SPECIAL = set(' []{}?=,\'|&<>')
    quoted = []
    for arg in args:
        if any(c in arg for c in _SPECIAL):
            # Escape any embedded double-quotes, then wrap in double-quotes
            arg = arg.replace('"', '\\"')
            quoted.append(f'"{arg}"')
        else:
            quoted.append(arg)
    cmd = "az " + " ".join(quoted) + " -o json"
    if verbose:
        print(f"  [az] {cmd}")
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120, shell=True)
        if result.returncode != 0:
            return None
        return json.loads(result.stdout) if result.stdout.strip() else None
    except (subprocess.TimeoutExpired, json.JSONDecodeError, FileNotFoundError):
        return None


# ── Regions input helpers ─────────────────────────────────────────────────────
def _parse_regions_string(raw: str) -> list[str]:
    """Split a comma-separated region string into a normalised list.

    Strips whitespace, lowercases, removes empties. Returns [] for empty input.
    """
    if not raw:
        return []
    return [r.strip().lower() for r in raw.split(",") if r and r.strip()]


def load_regions_file(path: str) -> list[str]:
    """Read a regions list from a text file.

    Format:
      • One region per line (e.g. `eastus`)
      • Lines starting with `#` are comments (in-line `#` also strips trailing comment)
      • Blank lines are ignored
      • Comma-separated tokens on a single line are also accepted
      • BOM (UTF-8 byte-order mark) at the start of file is handled
      • Duplicates are preserved here — dedup happens at merge time

    Raises FileNotFoundError if the file does not exist.
    """
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Regions file not found: {file_path}")

    out: list[str] = []
    with file_path.open("r", encoding="utf-8-sig") as fh:
        for line in fh:
            # Strip inline comments (anything after `#`) so users can annotate
            comment_idx = line.find("#")
            if comment_idx >= 0:
                line = line[:comment_idx]
            line = line.strip()
            if not line:
                continue
            for token in _parse_regions_string(line):
                out.append(token)
    return out


def _dedup_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def check_az_login() -> bool:
    """
    Return True if the user is logged in to Azure CLI.
    Uses `az account list` instead of `az account show` to avoid failures
    caused by inaccessible secondary tenants generating non-zero exit codes.
    """
    try:
        result = subprocess.run(
            "az account list -o json",
            capture_output=True, text=True, timeout=30, shell=True
        )
        # Succeed as long as stdout contains a non-empty JSON array,
        # regardless of return code (warnings about other tenants set rc=1)
        if result.stdout.strip():
            data = json.loads(result.stdout)
            return isinstance(data, list) and len(data) > 0
    except (subprocess.TimeoutExpired, json.JSONDecodeError, FileNotFoundError):
        pass
    return False


def get_all_regions(verbose: bool = False) -> list[dict]:
    """Return list of Azure public regions with {name, displayName}."""
    print("Fetching Azure region list...")
    regions = az(["account", "list-locations",
                   "--query", "[?metadata.regionType=='Physical'].{name:name,displayName:displayName}"],
                 verbose=verbose)
    return regions or []


# ── Provider availability ──────────────────────────────────────────────────────
_PROVIDER_CACHE: dict[str, list[str]] = {}


def get_provider_locations(provider: str, resource_type: str, verbose: bool = False) -> list[str]:
    """Return lowercase region names where a resource provider/type is available.
    Returns ['*'] if the service is globally available (e.g. Azure DNS zones)."""
    key = f"{provider}/{resource_type}"
    if key in _PROVIDER_CACHE:
        return _PROVIDER_CACHE[key]

    if verbose:
        print(f"  Querying provider locations: {key}")

    data = az(["provider", "show", "-n", provider,
               "--query", f"resourceTypes[?resourceType=='{resource_type}'].locations[]"],
              verbose=verbose)
    if data:
        # "global" means available everywhere — return a sentinel
        if any(loc.lower() == "global" for loc in data):
            normalised = ["*"]
        else:
            normalised = [loc.lower().replace(" ", "").replace("(", "").replace(")", "") for loc in data]
    else:
        normalised = []

    _PROVIDER_CACHE[key] = normalised
    return normalised


def _normalise(region_name: str) -> str:
    return region_name.lower().replace(" ", "").replace("(", "").replace(")", "")


# ── SSD v2 zone availability ───────────────────────────────────────────────────
def get_ssdv2_zone_info(verbose: bool = False) -> dict[str, list[str]]:
    """
    Query zone availability for Premium SSD v2 across all regions.
    Returns {region_name: [zone_list]} e.g. {"eastus": ["1","2","3"]}
    """
    data = az(["vm", "list-skus", "--resource-type", "disks",
               "--query", "[?name=='PremiumV2_LRS'].{region:locationInfo[0].location, zones:locationInfo[0].zones}"],
              verbose=verbose)
    result = {}
    if data:
        for item in data:
            if item.get("region"):
                result[item["region"].lower()] = item.get("zones") or []
    return result


# ── BOM reader ─────────────────────────────────────────────────────────────────
def load_bom(xlsx_path: str) -> list[dict]:
    """
    Reads the Catalog sheet to build the service lookup, then reads service
    names from column A of the BOM sheet (starting row 4).
    All listed services are treated as required.

    Catalog sheet columns: A=Service Name  B=Provider  C=Resource Type  D=Zone Check (Yes/No)
    BOM sheet column:      A=Service Name
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Catalog" not in wb.sheetnames:
        sys.exit(f"ERROR: No 'Catalog' sheet found in {xlsx_path}")
    if "BOM" not in wb.sheetnames:
        sys.exit(f"ERROR: No 'BOM' sheet found in {xlsx_path}")

    # Build catalog from spreadsheet
    catalog: dict[str, dict] = {}
    ws_cat = wb["Catalog"]
    for row in ws_cat.iter_rows(min_row=2, values_only=True):
        name  = row[0] if len(row) > 0 else None
        prov  = row[1] if len(row) > 1 else None
        rtype = row[2] if len(row) > 2 else None
        zone  = row[3] if len(row) > 3 else None
        if not name or not prov or not rtype:
            continue
        catalog[str(name).strip()] = {
            "provider":      str(prov).strip(),
            "resource_type": str(rtype).strip(),
            "zone_check":    str(zone or "No").strip().lower() == "yes",
        }

    # Read BOM
    ws = wb["BOM"]
    bom_rows: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        svc_name = row[0] if len(row) > 0 else None
        if not svc_name:
            continue
        svc_name = str(svc_name).strip()

        entry = catalog.get(svc_name)
        if not entry:
            print(f"  ⚠  '{svc_name}' not found in Catalog sheet — skipping.")
            continue

        bom_rows.append({
            "service":       svc_name,
            "provider":      entry["provider"],
            "resource_type": entry["resource_type"],
            "zone_check":    entry["zone_check"],
        })

    return bom_rows


def read_required_skus(xlsx_path: str) -> list[list[str]]:
    """Optional 'Required SKUs' sheet on the input BOM. Returns a list of
    [primary_family, primary_label, alt_family, alt_label, required_cores]
    rows so we can write them back through to the output workbook unchanged.
    Returns [] when the sheet is absent (older bom_template.xlsx files).

    The Required Cores column is optional; missing cells round-trip as ""."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return []
    if "Required SKUs" not in wb.sheetnames:
        return []
    ws = wb["Required SKUs"]
    rows = list(ws.iter_rows(values_only=True))
    out: list[list[str]] = []
    header_seen = False
    for raw in rows:
        if raw is None:
            continue
        cells = [("" if c is None else str(c)).strip() for c in raw]
        # Skip blanks and the header row itself.
        if not header_seen and any(c.lower() == "primary family" for c in cells):
            header_seen = True
            continue
        if not header_seen:
            continue
        cells = (cells + ["", "", "", "", ""])[:5]
        if not cells[0]:
            continue
        out.append(cells)
    return out


# ── Results writer ─────────────────────────────────────────────────────────────
def write_results(output_path: str, bom_rows: list[dict],
                  region_results: list[dict], check_time: str,
                  required_skus: list[list[str]] | None = None) -> None:
    """Write color-coded results to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Region Results"

    # ── Title ──
    t = ws.cell(row=1, column=1,
                value=f"Azure Region BOM Check  |  Generated: {check_time}")
    t.font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells(f"A1:{get_column_letter(len(bom_rows) + 4)}1")
    ws.row_dimensions[1].height = 24

    # ── Legend ──
    legend_row = 2
    for col, (colour, label) in enumerate([
        (FILL_GREEN, "All services available"),
        (FILL_RED,   "One or more services missing"),
    ], start=1):
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill = colour; c.font = Font(size=9); c.alignment = CENTER

    # ── Column headers: Region | DisplayName | Overall | <service1> … ──
    hdr_row = 3
    ws.cell(row=hdr_row, column=1, value="Region").fill = FILL_HDR
    ws.cell(row=hdr_row, column=1).font = FONT_HDR
    ws.cell(row=hdr_row, column=1).alignment = CENTER
    ws.cell(row=hdr_row, column=1).border = BORDER
    ws.column_dimensions["A"].width = 20

    ws.cell(row=hdr_row, column=2, value="Display Name").fill = FILL_HDR
    ws.cell(row=hdr_row, column=2).font = FONT_HDR
    ws.cell(row=hdr_row, column=2).alignment = CENTER
    ws.cell(row=hdr_row, column=2).border = BORDER
    ws.column_dimensions["B"].width = 22

    ws.cell(row=hdr_row, column=3, value="Overall Status").fill = FILL_HDR
    ws.cell(row=hdr_row, column=3).font = FONT_HDR
    ws.cell(row=hdr_row, column=3).alignment = CENTER
    ws.cell(row=hdr_row, column=3).border = BORDER
    ws.column_dimensions["C"].width = 18

    for col_offset, svc in enumerate(bom_rows, start=4):
        label = svc["service"]
        is_key = svc["service"] in ("Azure Automation", "Premium SSD v2 (PremiumV2_LRS)")
        c = ws.cell(row=hdr_row, column=col_offset, value=label)
        c.fill = FILL_KEY if is_key else FILL_HDR
        c.font = Font(bold=True, color="000000" if is_key else "FFFFFF", size=9)
        c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(col_offset)].width = 18
    ws.row_dimensions[hdr_row].height = 50

    # ── Data rows ──
    for r_offset, rr in enumerate(region_results, start=4):
        overall       = rr["overall"]
        overall_fill  = FILL_GREEN if overall == "PASS" else FILL_RED
        overall_label = "✅ SUPPORTED" if overall == "PASS" else "❌ UNSUPPORTED"

        ws.cell(row=r_offset, column=1, value=rr["region"]).border = BORDER
        ws.cell(row=r_offset, column=1).font = FONT_NORMAL
        ws.cell(row=r_offset, column=1).alignment = CENTER

        ws.cell(row=r_offset, column=2, value=rr.get("display_name", "")).border = BORDER
        ws.cell(row=r_offset, column=2).font = FONT_NORMAL
        ws.cell(row=r_offset, column=2).alignment = LEFT

        oc = ws.cell(row=r_offset, column=3, value=overall_label)
        oc.fill = overall_fill; oc.font = FONT_BOLD; oc.alignment = CENTER; oc.border = BORDER

        for col_offset, svc in enumerate(bom_rows, start=4):
            svc_name = svc["service"]
            svc_result = rr["services"].get(svc_name, {})
            available = svc_result.get("available")
            detail = svc_result.get("detail", "")

            if available is True:
                fill = FILL_GREEN
                label = f"✅ {detail}" if detail else "✅ Available"
            else:
                fill = FILL_RED
                label = f"❌ {detail}" if detail else "❌ Not available"

            c = ws.cell(row=r_offset, column=col_offset, value=label)
            c.fill = fill; c.font = Font(size=9); c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[r_offset].height = 40

    # Freeze headers
    ws.freeze_panes = f"A{hdr_row + 1}"

    # ── Notes sheet ──
    ws_notes = wb.create_sheet("Notes")
    notes = [
        ("Data source", "Resource provider availability is queried live from Azure Resource Manager "
         "via 'az provider show'. SSD v2 zone data is from 'az vm list-skus'."),
        ("Azure Automation", "Queried from Microsoft.Automation/automationAccounts provider locations. "
         "If not in the list, the service is not yet GA in that region."),
        ("Premium SSD v2", "Zone data from Microsoft.Compute disk SKU 'PremiumV2_LRS'. "
         "PASS requires 3 zones; WARN means available in <3 zones; FAIL means unavailable entirely."),
        ("VM SKUs", "Individual VM SKU availability (D8as v6 etc.) is NOT checked here. "
         "Use: az vm list-skus --location <region> --size Standard_D8as_v6"),
        ("Accuracy", "Provider location lists reflect current Azure public data for your subscription. "
         "Availability can change. For the latest info see: "
         "https://azure.microsoft.com/en-us/explore/global-infrastructure/products-by-region/"),
    ]
    ws_notes.column_dimensions["A"].width = 22
    ws_notes.column_dimensions["B"].width = 90
    for i, (label, detail) in enumerate(notes, start=1):
        ws_notes.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws_notes.cell(row=i, column=2, value=detail).font = Font(size=10)
        ws_notes.cell(row=i, column=2).alignment = LEFT
        ws_notes.row_dimensions[i].height = 36

    # ── Required SKUs sheet (passed through from the input BOM) ──
    # The Azure BOM Region Support Dashboard reads this sheet to drive ARM
    # queries. Only emit it when the input BOM actually had one so we don't
    # quietly invent SKU lists for users who haven't opted in.
    if required_skus:
        ws_skus = wb.create_sheet("Required SKUs")
        ws_skus.cell(row=1, column=1,
                     value="Required SKU families for this BOM "
                           "(consumed by the Azure BOM Region Support Dashboard).") \
               .font = Font(italic=True, size=9, color="666666")
        headers = ["Primary Family", "Primary Label", "Alt Family", "Alt Label",
                   "Required Cores"]
        for ci, h in enumerate(headers, start=1):
            c = ws_skus.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor="EEEEEE")
        for ri, row in enumerate(required_skus, start=3):
            for ci, val in enumerate(row, start=1):
                # Cores column: write as a number when possible so the
                # dashboard's downstream parser sees an int instead of a
                # string (cleaner export).
                if ci == 5 and val:
                    try:
                        val = int(val)
                    except (TypeError, ValueError):
                        try:
                            f = float(val)
                            if f.is_integer():
                                val = int(f)
                        except (TypeError, ValueError):
                            pass  # leave as-is; dashboard will reject
                ws_skus.cell(row=ri, column=ci, value=val)
        for col, w in zip("ABCDE", (28, 14, 28, 14, 14)):
            ws_skus.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")


# ── Main logic ─────────────────────────────────────────────────────────────────
def run_checks(bom_rows: list[dict], regions_to_check: list[dict],
               verbose: bool = False) -> list[dict]:
    """
    For each region, check availability of every BOM service.
    Returns list of region result dicts.
    """
    # Pre-fetch provider location lists for all non-zone services
    provider_services = [s for s in bom_rows if not s["zone_check"]]

    print(f"\nFetching service availability data...")
    provider_locations: dict[str, list[str]] = {}
    for svc in provider_services:
        if svc["provider"] and svc["resource_type"]:
            key = f"{svc['provider']}/{svc['resource_type']}"
            if key not in provider_locations:
                print(f"  Checking {svc['service']}...")
                locs = get_provider_locations(svc["provider"], svc["resource_type"], verbose)
                provider_locations[key] = locs

    # Pre-fetch SSDv2 zone info
    ssdv2_zones: dict[str, list[str]] = {}
    if any(s["zone_check"] for s in bom_rows):
        print(f"  Checking Premium SSD v2 (zone availability, this may take ~30 s)...")
        ssdv2_zones = get_ssdv2_zone_info(verbose)

    print(f"\nEvaluating {len(regions_to_check)} regions...")
    results = []

    for region_info in regions_to_check:
        region_name = region_info["name"]
        norm_name   = _normalise(region_name)
        display     = region_info.get("displayName", region_name)

        if verbose:
            print(f"  Checking {region_name} ({display})")

        services_result: dict[str, dict] = {}
        overall = "PASS"

        for svc in bom_rows:
            svc_name = svc["service"]

            if svc["zone_check"]:
                # Premium SSD v2 zone check
                zones = ssdv2_zones.get(region_name.lower()) or ssdv2_zones.get(norm_name, [])
                if not zones:
                    # Try alternate key formats
                    for k in ssdv2_zones:
                        if _normalise(k) == norm_name:
                            zones = ssdv2_zones[k]
                            break
                if len(zones) >= 3:
                    services_result[svc_name] = {"available": True, "detail": f"{len(zones)} zones"}
                else:
                    services_result[svc_name] = {"available": False,
                                                  "detail": f"only {len(zones)} zone(s)" if zones else "not available"}
                    overall = "FAIL"
            else:
                # Provider location check
                if not svc["provider"] or not svc["resource_type"]:
                    services_result[svc_name] = {"available": None, "detail": "no provider info"}
                    continue

                key = f"{svc['provider']}/{svc['resource_type']}"
                available_locs = provider_locations.get(key, [])

                # Match: global sentinel, exact name, or normalised name
                matched = ("*" in available_locs or
                           region_name.lower() in available_locs or
                           norm_name in available_locs or
                           any(_normalise(loc) == norm_name for loc in available_locs))

                if matched:
                    services_result[svc_name] = {"available": True, "detail": ""}
                else:
                    services_result[svc_name] = {"available": False, "detail": "not in provider list"}
                    overall = "FAIL"

        results.append({
            "region":       region_name,
            "display_name": display,
            "overall":      overall,
            "services":     services_result,
        })
        status_icon = "✅" if overall == "PASS" else "❌"
        print(f"  {status_icon} {region_name:25s} {display}")

    return results


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Check which Azure regions support your BOM.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--bom",     default="bom_template.xlsx",
                        help="BOM Excel file (default: bom_template.xlsx)")
    parser.add_argument("--output",  default="",
                        help="Output Excel file (default: region_results_<timestamp>.xlsx)")
    parser.add_argument("--regions", default="",
                        help="Comma-separated region names (e.g. eastus,westeurope). "
                             "Use the literal value `all` to check every Azure region.")
    parser.add_argument("--regions-file", default="",
                        help="Path to a text file containing regions to check (one per line, "
                             "`#` comments and blank lines OK). Combined with --regions if both given.")
    parser.add_argument("--verbose", action="store_true",
                        help="Print detailed progress including az CLI commands")
    args = parser.parse_args()

    # Resolve paths
    bom_path = Path(args.bom)
    if not bom_path.exists():
        sys.exit(f"ERROR: BOM file not found: {bom_path}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or f"region_results_{timestamp}.xlsx"

    # Check Azure CLI login
    print("Checking Azure CLI authentication...")
    if not check_az_login():
        sys.exit(
            "ERROR: Not logged in to Azure CLI.\n"
            "Run: az login\n"
            "Then re-run this script."
        )
    print("  ✅ Azure CLI authenticated\n")

    # Load BOM
    print(f"Loading BOM from: {bom_path}")
    bom_rows = load_bom(str(bom_path))
    print(f"  {len(bom_rows)} services loaded")

    # Determine regions to check — collect from --regions-file first, then --regions,
    # dedupe preserving order, then resolve display names against Azure.
    cli_regions = _parse_regions_string(args.regions)
    use_all_regions = any(r == "all" for r in cli_regions)
    cli_regions = [r for r in cli_regions if r != "all"]

    file_regions: list[str] = []
    if args.regions_file:
        try:
            file_regions = load_regions_file(args.regions_file)
            print(f"Loaded {len(file_regions)} region(s) from: {args.regions_file}")
        except FileNotFoundError as exc:
            sys.exit(f"ERROR: {exc}")

    all_regions = az(["account", "list-locations",
                      "--query", "[].{name:name,displayName:displayName}"],
                     args.verbose) or []
    region_map  = {r["name"].lower(): r for r in all_regions}

    if use_all_regions:
        # `--regions all` → check every Azure region returned by `az account list-locations`.
        # Any extra names from --regions or --regions-file are merged in (helpful for
        # newly-launched regions not yet in the CLI cache).
        region_names = _dedup_preserve_order(
            sorted(region_map.keys()) + file_regions + cli_regions
        )
    else:
        region_names = _dedup_preserve_order(file_regions + cli_regions)

    if not region_names:
        sys.exit(
            "ERROR: No target regions specified.\n"
            "  Use --regions, --regions-file, or `--regions all` to specify which\n"
            "  regions to check.\n"
            "  Examples:\n"
            "    python check_azure_regions.py --regions eastus,mexicocentral,westeurope\n"
            "    python check_azure_regions.py --regions-file regions.txt\n"
            "    python check_azure_regions.py --regions all"
        )

    regions_to_check = [region_map.get(n, {"name": n, "displayName": n}) for n in region_names]
    print(f"Checking {len(regions_to_check)} region(s)...\n")

    if not regions_to_check:
        sys.exit("ERROR: No regions to check.")

    # Run checks
    check_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    results = run_checks(bom_rows, regions_to_check, verbose=args.verbose)

    # Pass through the optional Required SKUs sheet from the input BOM so the
    # output region_results_*.xlsx is a complete per-customer config the
    # Azure BOM Region Support Dashboard can consume directly.
    required_skus = read_required_skus(args.bom)

    # Write output
    write_results(output_path, bom_rows, results, check_time,
                  required_skus=required_skus)

    # Print summary
    passed = sum(1 for r in results if r["overall"] == "PASS")
    failed = sum(1 for r in results if r["overall"] == "FAIL")
    print(f"\n{'─'*55}")
    print(f"Summary: {len(results)} regions checked")
    print(f"  ✅ SUPPORTED   : {passed}")
    print(f"  ❌ UNSUPPORTED : {failed}")
    print(f"{'─'*55}")

    if passed:
        supported = [r["region"] for r in results if r["overall"] == "PASS"]
        print(f"\nSupported regions: {', '.join(supported)}")




if __name__ == "__main__":
    main()
