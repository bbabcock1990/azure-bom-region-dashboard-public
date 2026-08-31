"""Compile orchestrator for ARM-backed BOM + SKU analysis."""
from __future__ import annotations

import io
import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timezone
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl

from . import (
    activity_log,
    arm_sku_availability,
    arm_skus,
    quota_groups,
    run_progress,
    storage,
)
from .pipeline import model as pipeline_model
from .pipeline import sources as pipeline_sources

log = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
ENGINE_VERSION = "swa-1.3.0"

REQUIRED_SKUS_SHEET = "Required SKUs"

# Azure VM SKU family identifiers are ASCII-alphanumeric (e.g.
# standardDav6Family). We use this both to detect end-of-data in the BOM
# Required SKUs sheet (so notes/bullets below the data don't get parsed
# as SKUs even without a blank separator row) and to give a clear
# validation error when a user pastes garbage into the override textarea.
_FAMILY_ID_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_-]*$")


def _parse_required_cores(cell, *, where: str):
    """Normalize a 'Required Cores' value coming from any source (Excel cell,
    override textarea, JSON API). Returns:
      - ``None`` when the value is blank/absent (no requirement).
      - A positive ``int`` otherwise.
    Raises ``CompileError`` with a stable code on bad input.

    Accepted forms:
      * ``None`` / ``""`` / whitespace      → ``None``
      * ``int`` (e.g. 100)                  → ``100``
      * ``float`` with zero fractional part → coerced to ``int``
      * numeric ``str`` (``"100"``, ``"100.0"``) → parsed
    Rejected:
      * ``0`` or negative (use blank to mean 'no requirement')
      * ``bool`` (``True``/``False`` — easy to slip through since
        ``isinstance(True, int)`` is True in Python)
      * floats with non-zero fractional part (cores are whole vCPUs)
      * NaN / inf
      * any string with non-numeric content (units, etc.)
    """
    # Blank / missing
    if cell is None:
        return None
    if isinstance(cell, str):
        s = cell.strip()
        if not s:
            return None
        # Try int first so "100" stays exact; fall back to float for "100.0".
        try:
            return _coerce_positive_int(int(s), where=where)
        except ValueError:
            pass
        try:
            f = float(s)
        except ValueError:
            raise CompileError(
                "bad_required_cores",
                f"{where}: required_cores {cell!r} is not a number.",
                400,
            )
        return _coerce_positive_int(f, where=where)
    # bool is a subclass of int in Python — reject explicitly so True/False
    # don't slip through as 1/0.
    if isinstance(cell, bool):
        raise CompileError(
            "bad_required_cores",
            f"{where}: required_cores must be a positive integer, not a boolean.",
            400,
        )
    if isinstance(cell, (int, float)):
        return _coerce_positive_int(cell, where=where)
    raise CompileError(
        "bad_required_cores",
        f"{where}: required_cores has unsupported type {type(cell).__name__}.",
        400,
    )


def _coerce_positive_int(value, *, where: str) -> int:
    """Helper for _parse_required_cores. Accepts a numeric value (int/float)
    and returns a positive int, or raises CompileError. Floats with a
    non-zero fractional part are rejected (cores are whole vCPUs)."""
    import math
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            raise CompileError(
                "bad_required_cores",
                f"{where}: required_cores must be a finite number.",
                400,
            )
        if not value.is_integer():
            raise CompileError(
                "bad_required_cores",
                (f"{where}: required_cores must be a whole number "
                 f"(got {value!r})."),
                400,
            )
        value = int(value)
    if not isinstance(value, int) or isinstance(value, bool):
        raise CompileError(
            "bad_required_cores",
            f"{where}: required_cores must be an integer (got {value!r}).",
            400,
        )
    if value < 1:
        raise CompileError(
            "bad_required_cores",
            (f"{where}: required_cores must be a positive integer "
             f"(got {value}). Leave blank to mean 'no requirement'."),
            400,
        )
    return value


class CompileError(Exception):
    def __init__(self, code: str, message: str, status: int = 400):
        super().__init__(message)
        self.code = code
        self.message = message
        self.status = status


def _normalize_subscriptions(
    *,
    subscription_id: Optional[str],
    arm_token: Optional[str],
    subscriptions: Optional[List[Dict]],
) -> List[Dict]:
    if subscriptions is None:
        if not (subscription_id or "").strip():
            raise CompileError("missing_subscription", "subscription_id is required.", 400)
        if not arm_token:
            raise CompileError("missing_token", "ARM token is required.", 400)
        return [{
            "subscription_id": str(subscription_id).strip().lower(),
            "arm_token": arm_token,
            "status": "ok",
        }]
    if not isinstance(subscriptions, list) or not subscriptions:
        raise CompileError(
            "missing_subscription",
            "subscriptions must contain at least one subscription.",
            400,
        )
    out: List[Dict] = []
    seen = set()
    for i, item in enumerate(subscriptions):
        if not isinstance(item, dict):
            raise CompileError(
                "bad_subscription",
                f"subscriptions[{i}] must be an object.",
                400,
            )
        sub_id = str(item.get("subscription_id") or "").strip().lower()
        if not sub_id:
            raise CompileError(
                "bad_subscription",
                f"subscriptions[{i}].subscription_id is required.",
                400,
            )
        if sub_id in seen:
            continue
        seen.add(sub_id)
        token = item.get("arm_token")
        status = str(item.get("status") or ("ok" if token else "no_access")).strip() or "unknown"
        out.append({
            "subscription_id": sub_id,
            "arm_token": token,
            "status": status,
            "error": item.get("error"),
            "role": item.get("role") or ("target" if i == 0 else "auxiliary"),
        })
    return out


def _canonical_family_map(families: Iterable[str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for family in families:
        if family and str(family).strip():
            out.setdefault(str(family).strip().lower(), str(family).strip())
    return out


def _placeholder_sku_records(
    *,
    subscription_id: str,
    regions: Iterable[str],
    families: Iterable[str],
    reason: str,
    status: str,
) -> List[Dict]:
    family_canonical = _canonical_family_map(families)
    want_regions = {
        arm_sku_availability.arm_skus._normalize_region(r)  # noqa: SLF001
        for r in regions if r and str(r).strip()
    }
    rows: List[Dict] = []
    for region in sorted(want_regions):
        for fam_lower, fam_canonical in sorted(family_canonical.items()):
            rows.append({
                "region": region,
                "family": fam_canonical,
                "display": arm_sku_availability._friendly_family(fam_canonical),  # noqa: SLF001
                "zones": [False, False, False],
                "sub_restricted": True,
                "sub_restriction_raw": reason,
                "arm_provenance": {
                    "available": False,
                    "region_restricted": True,
                    "restricted_zones": ["1", "2", "3"],
                    "available_zones": [],
                    "notes": [reason],
                    "status": status,
                    "subscription_id": subscription_id,
                },
            })
    return rows


def _index_sku_records(rows: List[Dict]) -> Dict[Tuple[str, str], Dict]:
    out: Dict[Tuple[str, str], Dict] = {}
    for row in rows or []:
        region = arm_sku_availability.arm_skus._normalize_region(row.get("region"))  # noqa: SLF001
        family = str(row.get("family") or "").strip().lower()
        if region and family:
            out[(region, family)] = row
    return out


def _quota_num(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except Exception:
        return None


def _quota_headroom(limit, usage) -> Optional[float]:
    limit_num = _quota_num(limit)
    usage_num = _quota_num(usage)
    if limit_num is None or usage_num is None:
        return None
    return limit_num - usage_num


def _normalize_quota_value(value):
    if value is None:
        return None
    if float(value).is_integer():
        return int(value)
    return round(float(value), 2)


def _subscription_quota_region(
    subscription_quota: Optional[Dict],
    region_short: str,
) -> Optional[Dict]:
    if not subscription_quota:
        return None
    return (subscription_quota.get("regions") or {}).get(region_short.lower())


def _lookup_family_row(family_rows, family_id: str) -> Optional[Dict]:
    family_key = str(family_id or "").strip().lower()
    if not family_key or not isinstance(family_rows, dict):
        return None
    for name, row in family_rows.items():
        if str(name or "").strip().lower() == family_key:
            return row
    return None


def _best_subscription_quota_match(
    *,
    subscription_quota: Optional[Dict],
    region_short: str,
    family_ids: List[str],
    required_cores: int,
    subscription_id: str,
) -> Dict:
    region_info = _subscription_quota_region(subscription_quota, region_short) or {}
    families = region_info.get("families") or {}
    candidates: List[Dict] = []
    for order, family_id in enumerate(family_ids):
        row = _lookup_family_row(families, family_id)
        if not row:
            continue
        limit = _normalize_quota_value(_quota_num(row.get("limit")))
        usage = _normalize_quota_value(_quota_num(row.get("usage")))
        headroom = _normalize_quota_value(
            _quota_headroom(row.get("limit"), row.get("usage"))
            if row.get("headroom") is None else _quota_num(row.get("headroom"))
        )
        candidates.append({
            "subscription_id": subscription_id,
            "family": family_id,
            "limit": limit,
            "usage": usage,
            "headroom": headroom,
            "sufficient": headroom is not None and headroom >= required_cores,
            "order": order,
        })
    best = None
    if candidates:
        best = sorted(
            candidates,
            key=lambda item: (
                0 if item["sufficient"] else 1,
                -(item["headroom"] if item["headroom"] is not None else -1e15),
                item["order"],
            ),
        )[0]
    total_regional = (region_info or {}).get("total_regional")
    return {
        "subscription_id": subscription_id,
        "status": (
            (region_info or {}).get("status")
            or ((subscription_quota or {}).get("status"))
            or "unknown"
        ) if best is None else "ok",
        "error": (
            (region_info or {}).get("error")
            or (subscription_quota or {}).get("error")
        ),
        "family": best.get("family") if best else (family_ids[0] if family_ids else None),
        "limit": best.get("limit") if best else None,
        "usage": best.get("usage") if best else None,
        "headroom": best.get("headroom") if best else None,
        "sufficient": bool(best and best.get("sufficient")),
        "total_regional": total_regional,
    }


def _best_quota_group_match(
    *,
    quota_result: Optional[Dict],
    region_short: str,
    family_ids: List[str],
    shortfall: int,
    subscription_id: str,
) -> Dict:
    groups = [
        g for g in ((quota_result or {}).get("groups") or [])
        if str(g.get("region") or "").strip().lower() == region_short.lower()
    ]
    candidates: List[Dict] = []
    for group in groups:
        for order, family_id in enumerate(family_ids):
            for fam in (group.get("families") or []):
                if str(fam.get("family") or "").strip().lower() != family_id.lower():
                    continue
                limit = _normalize_quota_value(_quota_num(fam.get("limit")))
                usage = _normalize_quota_value(_quota_num(fam.get("usage")))
                headroom = _normalize_quota_value(_quota_headroom(fam.get("limit"), fam.get("usage")))
                candidates.append({
                    "subscription_id": subscription_id,
                    "group": group.get("name"),
                    "family": fam.get("family") or family_id,
                    "limit": limit,
                    "usage": usage,
                    "headroom": headroom,
                    "sufficient": headroom is not None and headroom >= shortfall,
                    "order": order,
                })
    best = None
    if candidates:
        best = sorted(
            candidates,
            key=lambda item: (
                0 if item["sufficient"] else 1,
                -(item["headroom"] if item["headroom"] is not None else -1e15),
                item["order"],
            ),
        )[0]
    return {
        "subscription_id": subscription_id,
        "status": (quota_result or {}).get("status") or "unknown",
        "error": (quota_result or {}).get("error"),
        "available": bool(best),
        "group": best.get("group") if best else None,
        "family": best.get("family") if best else (family_ids[0] if family_ids else None),
        "limit": best.get("limit") if best else None,
        "usage": best.get("usage") if best else None,
        "headroom": best.get("headroom") if best else None,
        "shortfall": shortfall,
        "sufficient": bool(best and best.get("sufficient")),
    }


def _evaluate_subscription_requirement(
    *,
    region_short: str,
    sub_id: str,
    sub_result: Dict,
    required: Dict,
) -> Dict:
    required_cores = int(required.get("required_cores") or 0)
    family_ids = [required.get("primary_family")]
    if required.get("alt_family"):
        family_ids.append(required.get("alt_family"))
    family_ids = [f for f in family_ids if f]
    tier1 = _best_subscription_quota_match(
        subscription_quota=sub_result.get("subscription_quota"),
        region_short=region_short,
        family_ids=family_ids,
        required_cores=required_cores,
        subscription_id=sub_id,
    )
    base_headroom = tier1.get("headroom")
    shortfall = required_cores
    if base_headroom is not None:
        shortfall = max(0, required_cores - int(float(base_headroom)))
    tier2 = _best_quota_group_match(
        quota_result=sub_result.get("quota_groups"),
        region_short=region_short,
        family_ids=family_ids,
        shortfall=shortfall,
        subscription_id=sub_id,
    )

    if tier1.get("sufficient"):
        overall_status = "sufficient_sub"
    elif tier2.get("sufficient"):
        overall_status = "sufficient_group"
    else:
        sub_region = _subscription_quota_region(sub_result.get("subscription_quota"), region_short) or {}
        sub_known = tier1.get("headroom") is not None
        group_known = tier2.get("headroom") is not None
        if sub_region.get("status") not in (None, "ok") and not group_known:
            overall_status = "unknown"
        elif sub_known or group_known or tier2.get("status") in ("no_quota_group", "not_available"):
            overall_status = "insufficient"
        else:
            overall_status = "unknown"
    return {
        "overall_status": overall_status,
        "tier1_sub_quota": tier1,
        "tier2_quota_group": tier2,
    }


def _summarize_subscription_quota_for_region(
    *,
    region_short: str,
    required_families: List[Dict],
    sub_id: str,
    sub_result: Dict,
    sub_status: str,
    sub_error: Optional[str],
) -> Dict:
    quota_result = sub_result.get("quota_groups") or {}
    summary = {
        "subscription_id": sub_id,
        "status": "unknown",
        "error": sub_error,
        "has_quota_groups": bool(quota_result.get("has_quota_groups")),
        "groups": [
            g for g in (quota_result.get("groups") or [])
            if str(g.get("region") or "").strip().lower() == region_short.lower()
        ],
        "subscription_quota": _subscription_quota_region(
            sub_result.get("subscription_quota"), region_short,
        ),
    }
    if sub_status != "ok":
        summary["status"] = sub_status or "unknown"
        return summary
    reqs = [r for r in (required_families or []) if r.get("required_cores")]
    if not reqs:
        return summary

    any_sufficient = False
    any_group = False
    any_insufficient = False
    any_unknown = False
    for req in reqs:
        evaluated = _evaluate_subscription_requirement(
            region_short=region_short,
            sub_id=sub_id,
            sub_result=sub_result,
            required=req,
        )
        status = evaluated["overall_status"]
        if status == "sufficient_sub":
            any_sufficient = True
        elif status == "sufficient_group":
            any_sufficient = True
            any_group = True
        elif status == "insufficient":
            any_insufficient = True
        else:
            any_unknown = True

    if any_insufficient and any_sufficient:
        summary["status"] = "partial"
    elif any_insufficient:
        summary["status"] = "insufficient"
    elif any_unknown and not any_sufficient:
        summary["status"] = "unknown"
    elif any_unknown:
        summary["status"] = "unknown"
    elif any_group:
        summary["status"] = "sufficient_group"
    elif any_sufficient:
        summary["status"] = "sufficient_sub"
    elif quota_result.get("status") in ("no_quota_group", "not_available"):
        summary["status"] = quota_result.get("status")
    return summary


def _evaluate_quota_tiers(
    *,
    per_sub_results: Dict[str, Dict],
    required_families: List[Dict],
    regions: List[str],
    target_subscription_id: Optional[str],
) -> Dict:
    target_sub_id = target_subscription_id or next(iter(per_sub_results.keys()), None)
    reqs = [r for r in (required_families or []) if r.get("required_cores")]
    region_results: Dict[str, Dict] = {}
    for region_short in sorted({
        arm_sku_availability.arm_skus._normalize_region(r)  # noqa: SLF001
        for r in regions if r and str(r).strip()
    }):
        family_results: Dict[str, Dict] = {}
        family_statuses: List[str] = []
        for req in reqs:
            key = str(req.get("primary_family") or "").strip().lower()
            if not key:
                continue
            need = int(req.get("required_cores") or 0)
            target_result = per_sub_results.get(target_sub_id or "") or {}
            target_eval = _evaluate_subscription_requirement(
                region_short=region_short,
                sub_id=target_sub_id or "",
                sub_result=target_result,
                required=req,
            )
            shortfall = target_eval["tier2_quota_group"].get("shortfall") or need
            cross_sub: List[Dict] = []
            family_ids = [req.get("primary_family")]
            if req.get("alt_family"):
                family_ids.append(req.get("alt_family"))
            family_ids = [f for f in family_ids if f]
            for sub_id, sub_result in per_sub_results.items():
                if not sub_id or sub_id == target_sub_id:
                    continue
                donor = _best_subscription_quota_match(
                    subscription_quota=sub_result.get("subscription_quota"),
                    region_short=region_short,
                    family_ids=family_ids,
                    required_cores=need,
                    subscription_id=sub_id,
                )
                if donor.get("headroom") is None or donor.get("headroom") <= 0:
                    continue
                donor["sufficient"] = donor["headroom"] >= shortfall
                cross_sub.append(donor)
            cross_sub.sort(
                key=lambda item: -float(item.get("headroom") or 0),
            )

            satisfied_by = None
            if target_eval["overall_status"] == "sufficient_sub":
                overall_status = "sufficient"
                satisfied_by = "subscription"
            elif target_eval["overall_status"] == "sufficient_group":
                overall_status = "sufficient"
                satisfied_by = "quota_group"
            elif target_eval["overall_status"] == "insufficient":
                overall_status = "insufficient"
            else:
                overall_status = "unknown"

            family_results[key] = {
                "family": req.get("primary_family"),
                "label": req.get("primary_label") or req.get("primary_family"),
                "alt_family": req.get("alt_family"),
                "alt_label": req.get("alt_label"),
                "required_cores": need,
                "tier1_sub_quota": target_eval["tier1_sub_quota"],
                "tier2_quota_group": target_eval["tier2_quota_group"],
                "tier3_cross_sub": cross_sub,
                "overall_status": overall_status,
                "satisfied_by": satisfied_by,
            }
            family_statuses.append(overall_status)

        if family_statuses and all(status == "sufficient" for status in family_statuses):
            region_status = "sufficient"
        elif any(status == "insufficient" for status in family_statuses):
            region_status = "partial" if any(status == "sufficient" for status in family_statuses) else "insufficient"
        else:
            region_status = "unknown"
        region_results[region_short] = {
            "subscription_id": target_sub_id,
            "status": region_status,
            "families": family_results,
        }
    return region_results


def _round_up_to_nearest_10(value: int) -> int:
    value = int(value or 0)
    if value <= 0:
        return 0
    return ((value + 9) // 10) * 10


def _build_quota_remediation_plan(
    *,
    per_sub_results: Dict[str, Dict],
    quota_tiers: Dict,
    regions: List[Dict],
    required_families: List[Dict],
) -> List[Dict]:
    """Generate quota increase requests needed to make quota-limited regions viable."""
    reqs = [r for r in (required_families or []) if r.get("required_cores")]
    region_names = {
        str(region.get("short") or "").strip().lower(): (
            region.get("name")
            or region.get("display_name")
            or region.get("short")
        )
        for region in (regions or [])
        if isinstance(region, dict) and str(region.get("short") or "").strip()
    }
    remediation: List[Dict] = []

    for region_short, region_quota in (quota_tiers or {}).items():
        if (region_quota or {}).get("status") != "insufficient":
            continue
        target_sub_id = str((region_quota or {}).get("subscription_id") or "").strip().lower()
        if not target_sub_id:
            target_sub_id = str(next(iter(per_sub_results.keys()), "") or "").strip().lower()
        target_sub_result = per_sub_results.get(target_sub_id) or {}
        family_results = (region_quota or {}).get("families") or {}
        for req in reqs:
            primary_family = str(req.get("primary_family") or "").strip()
            primary_key = primary_family.lower()
            if not primary_key:
                continue
            family_quota = family_results.get(primary_key) or {}
            if family_quota.get("overall_status") != "insufficient":
                continue

            required_cores = int(
                family_quota.get("required_cores") or req.get("required_cores") or 0
            )
            if required_cores <= 0:
                continue

            family_ids = [primary_family]
            if req.get("alt_family"):
                family_ids.append(req["alt_family"])

            best_match = _best_subscription_quota_match(
                subscription_quota=target_sub_result.get("subscription_quota"),
                region_short=region_short,
                family_ids=family_ids,
                required_cores=required_cores,
                subscription_id=target_sub_id,
            )
            if best_match is None:
                continue
            if best_match.get("headroom") is None:
                continue

            chosen_family = str(best_match.get("family") or primary_family).strip()
            chosen_family_lower = chosen_family.lower()
            family_label = (
                req.get("alt_label")
                if chosen_family_lower
                and chosen_family_lower == str(req.get("alt_family") or "").strip().lower()
                else req.get("primary_label")
            ) or family_quota.get("label") or chosen_family

            headroom = int(float(best_match.get("headroom") or 0))
            current_usage = _normalize_quota_value(_quota_num(best_match.get("usage")))
            current_limit = _normalize_quota_value(_quota_num(best_match.get("limit")))
            increase_needed = max(0, required_cores - headroom)
            recommended_base = (
                float(current_usage) + required_cores
                if current_usage is not None
                else (
                    float(current_limit) + increase_needed
                    if current_limit is not None
                    else required_cores
                )
            )
            new_limit_recommended = _round_up_to_nearest_10(int(recommended_base))

            remediation.append({
                "region": region_short,
                "region_display": region_names.get(region_short) or region_short,
                "subscription_id": best_match.get("subscription_id"),
                "family": chosen_family,
                "family_label": family_label,
                "current_limit": current_limit,
                "current_usage": current_usage,
                "required_cores": required_cores,
                "increase_needed": increase_needed,
                "new_limit_recommended": new_limit_recommended,
                "priority": "critical" if headroom <= 0 else "high",
            })

    priority_order = {"critical": 0, "high": 1}
    remediation.sort(
        key=lambda item: (
            priority_order.get(str(item.get("priority") or "").lower(), 99),
            str(item.get("region_display") or item.get("region") or "").lower(),
            str(item.get("family_label") or item.get("family") or "").lower(),
            str(item.get("subscription_id") or "").lower(),
        )
    )
    return remediation


def _apply_quota_group_status(
    *,
    snapshot_regions: List[Dict],
    required_families: List[Dict],
    per_sub_results: Dict[str, Dict],
    per_sub_status: Dict[str, Dict],
    tiered_results: Dict[str, Dict],
) -> None:
    for region in snapshot_regions:
        region_short = str(region.get("short") or "").strip().lower()
        if not region_short:
            region["quota_status"] = "unknown"
            region["quota_tiers"] = {"status": "unknown", "families": {}}
            region["quota_subscriptions"] = []
            continue
        sub_summaries: List[Dict] = []
        for sub_id, status_info in per_sub_status.items():
            summary = _summarize_subscription_quota_for_region(
                region_short=region_short,
                required_families=required_families,
                sub_id=sub_id,
                sub_result=per_sub_results.get(sub_id) or {},
                sub_status=status_info.get("status") or "unknown",
                sub_error=status_info.get("error"),
            )
            sub_summaries.append(summary)
        tiered = tiered_results.get(region_short) or {"status": "unknown", "families": {}}
        region["quota_status"] = tiered.get("status") or "unknown"
        region["quota_tiers"] = tiered
        region["quota_subscriptions"] = sub_summaries


def _append_unique_message(messages: List[str], message: Optional[str]) -> None:
    text = str(message or "").strip()
    if text and text not in messages:
        messages.append(text)


def _append_unique_blocker(
    blockers: List[Dict],
    *,
    blocker_type: str,
    message: str,
    severity: str,
) -> None:
    entry = {
        "type": blocker_type,
        "message": str(message).strip(),
        "severity": severity,
    }
    if entry["message"] and entry not in blockers:
        blockers.append(entry)


def _zones_label(zones: Optional[List[bool]]) -> str:
    if not zones:
        return "0 of 3 zones"
    present = [str(i + 1) for i, ok in enumerate(zones[:3]) if ok]
    if len(present) == 3:
        return "all 3 zones"
    if not present:
        return "0 of 3 zones"
    return f"{len(present)} of 3 zones (AZ {'/'.join(present)})"


def _merge_family_zone_health(
    *,
    region_short: str,
    family_ids: List[str],
    per_sub_results: Dict[str, Dict],
) -> Dict:
    zones = [False, False, False]
    any_access = False
    found = False
    notes: List[str] = []
    for info in per_sub_results.values():
        if (info.get("status") or "").strip().lower() != "ok":
            continue
        any_access = True
        sku_index = info.get("sku_index") or _index_sku_records(
            info.get("sku_records") or [],
        )
        for family_id in family_ids:
            fam_key = str(family_id or "").strip().lower()
            if not fam_key:
                continue
            row = sku_index.get((region_short, fam_key))
            if not row:
                continue
            found = True
            row_zones = list(row.get("zones") or [False, False, False])
            for i in range(3):
                zones[i] = zones[i] or bool(row_zones[i])
            note = row.get("sub_restriction_raw")
            if note:
                _append_unique_message(notes, str(note))
    return {
        "zones": zones if found else None,
        "any_access": any_access,
        "found": found,
        "notes": notes,
    }


def _quota_family_satisfaction_source(family_info: Dict) -> Optional[Dict]:
    satisfied_by = family_info.get("satisfied_by")
    if satisfied_by == "subscription":
        return family_info.get("tier1_sub_quota")
    if satisfied_by == "quota_group":
        return family_info.get("tier2_quota_group")
    return None


def _format_quota_number(value) -> str:
    number = _quota_num(value)
    if number is None:
        return "unknown"
    return str(int(number)) if float(number).is_integer() else str(round(number, 2))


def _quota_tight_message(label: str, required_cores: int, source: Optional[Dict]) -> Optional[str]:
    if not isinstance(source, dict):
        return None
    limit = _quota_num(source.get("limit"))
    usage = _quota_num(source.get("usage"))
    headroom = _quota_num(source.get("headroom"))
    if limit is not None and usage is not None and limit > 0:
        pct_used = (usage / limit) * 100
        if pct_used >= 80:
            return f"Quota tight for {label} ({pct_used:.0f}% used)"
    if headroom is not None and required_cores > 0 and headroom <= max(required_cores * 1.25, 10):
        return (
            f"Quota tight for {label} "
            f"({ _format_quota_number(headroom) } vCPU headroom for {required_cores} required)"
        )
    return None


# Documentation deep-links surfaced with recommendations.
_ODCR_DOC_URL = "https://learn.microsoft.com/azure/virtual-machines/capacity-reservation-overview"
_ZONE_ACCESS_DOC_URL = (
    "https://learn.microsoft.com/azure/virtual-machines/"
    "zonal-enablement-request-for-restricted-vm-series"
)
_QUOTA_DOC_URL = (
    "https://learn.microsoft.com/azure/quotas/quickstart-increase-quota-portal"
)


def _build_recommendations(
    *,
    blockers: List[Dict],
    region: Dict,
    region_display: str,
    fallback_used: bool,
    restricted_notes: List[str],
    constrained_labels: List[str],
    quota_required: bool,
) -> List[Dict]:
    """Map blocker/restriction signals to actionable mitigations.

    The headline recommendation is On-Demand Capacity Reservation (ODCR) for
    capacity-constrained or restricted regions, where reserving capacity in a
    specific zone de-risks allocation failures. Other recommendations point the
    operator at the correct support ticket (zonal access / quota increase),
    the fallback SKU, or an alternate region.
    """
    types = {str(b.get("type") or "") for b in blockers}
    has_zone_gap = "zone_gap" in types
    has_sku_unavailable = "sku_unavailable" in types
    has_quota = "quota_insufficient" in types
    has_missing_service = "missing_service" in types
    has_no_access = "no_access" in types
    is_restricted = bool(restricted_notes)
    capacity_constrained = has_zone_gap or has_sku_unavailable or is_restricted

    labels = ", ".join(constrained_labels) if constrained_labels else "the required VM series"
    recs: List[Dict] = []

    # 1) On-Demand Capacity Reservation — the headline recommendation for
    #    restricted / capacity-constrained regions.
    if capacity_constrained and not (has_no_access and not is_restricted):
        detail = (
            f"{region_display} shows capacity pressure for {labels}. "
            "Consider an On-Demand Capacity Reservation (ODCR) to guarantee "
            "allocation for a specific VM size in a specific availability zone "
            "before you deploy — this is most valuable in restricted or "
            "high-demand regions where allocation can fail intermittently. "
            "Note: an ODCR consumes vCPU quota for the reserved cores (raise "
            "quota first if headroom is tight) and requires the SKU to be "
            "offered to your subscription in that zone (request zonal access "
            "first if the zone is restricted)."
        )
        recs.append({
            "type": "odcr",
            "title": "Reserve capacity with On-Demand Capacity Reservations",
            "detail": detail,
            "priority": "high" if (has_zone_gap or has_sku_unavailable) else "medium",
            "doc_url": _ODCR_DOC_URL,
        })

    # 2) Zonal access ticket — restricted zones need whitelisting.
    if has_zone_gap and is_restricted:
        recs.append({
            "type": "zonal_access",
            "title": "Request zonal access for the restricted zone(s)",
            "detail": (
                "One or more zones are restricted for this VM series in "
                f"{region_display}. File a zonal access (SKU restriction) "
                "request from the Support tab to have the zones whitelisted "
                "for your subscription."
            ),
            "priority": "high",
            "doc_url": _ZONE_ACCESS_DOC_URL,
            "ticket_kind": "technical",
        })

    # 3) Quota increase ticket.
    if has_quota:
        recs.append({
            "type": "quota_increase",
            "title": "Raise the vCPU quota for the required series",
            "detail": (
                "Requested cores exceed the current vCPU limit. Open a quota "
                "increase from the Support tab — it pre-fills the exact "
                "shortfall (needed minus current usage) as the new limit."
            ),
            "priority": "high",
            "doc_url": _QUOTA_DOC_URL,
            "ticket_kind": "quota",
        })

    # 4) Fallback SKU already covers the region.
    if fallback_used:
        recs.append({
            "type": "fallback_sku",
            "title": "Standardize on the fallback SKU here",
            "detail": (
                "The primary series is thin in this region but the fallback "
                "series covers all zones. Plan to deploy the fallback SKU in "
                f"{region_display} (keep the primary where it is fully "
                "available) to avoid allocation gaps."
            ),
            "priority": "medium",
        })

    # 5) Alternate region when capacity/service can't be secured here.
    if has_missing_service or (has_sku_unavailable and not fallback_used):
        recs.append({
            "type": "alt_region",
            "title": "Evaluate an alternate region",
            "detail": (
                "If capacity or a required service can't be secured in "
                f"{region_display}, compare the nearest BOM-approved region "
                "with a Ready verdict as a deployment target or overflow."
            ),
            "priority": "medium",
        })

    # 6) Access / permissions so automated validation can run.
    if has_no_access:
        recs.append({
            "type": "grant_access",
            "title": "Grant Reader so ARM checks can validate this region",
            "detail": (
                "Automated SKU/quota validation could not run for the target "
                "subscription. Grant at least Reader on the subscription (or "
                "validate the region manually) to get an authoritative verdict."
            ),
            "priority": "high",
        })

    return recs


def _compute_deployment_verdict(
    region: Dict,
    *,
    snapshot_meta: Dict,
    per_sub_results: Dict[str, Dict],
) -> Dict:
    region_short = arm_sku_availability.arm_skus._normalize_region(  # noqa: SLF001
        region.get("short"),
    )
    required_families = list(snapshot_meta.get("skus_resolved") or [])
    quota_required = any(req.get("required_cores") for req in required_families)
    reasons: List[str] = []
    constraints: List[str] = []
    blockers: List[Dict] = []

    missing_services = list(region.get("missing_services") or [])
    if missing_services:
        for item in missing_services:
            service = str(item.get("service") or "Required service").strip()
            detail = str(item.get("detail") or "Not available").strip()
            _append_unique_blocker(
                blockers,
                blocker_type="missing_service",
                message=f"{service} not available ({detail})",
                severity="critical",
            )
    else:
        _append_unique_message(reasons, "All required services available")

    registration_required = list(region.get("registration_required") or [])
    for item in registration_required:
        service = str(item.get("service") or "Required service").strip()
        provider = str(item.get("provider") or "").strip()
        prov_note = f" — register {provider}" if provider else ""
        _append_unique_message(
            constraints,
            f"{service} requires resource-provider registration{prov_note}",
        )

    target_sub_id = (
        snapshot_meta.get("target_subscription_id")
        or snapshot_meta.get("subscription_id")
    )
    per_sub_status = snapshot_meta.get("per_sub_status") or {}
    target_status = per_sub_status.get(target_sub_id) or {}
    any_arm_access = any(
        (info.get("status") or "").strip().lower() == "ok"
        for info in per_sub_results.values()
    )
    if snapshot_meta.get("mode") == "global_unscoped":
        _append_unique_blocker(
            blockers,
            blocker_type="no_access",
            message=(
                snapshot_meta.get("mode_note")
                or target_status.get("error")
                or "Target subscription ARM access unavailable; validate this region manually."
            ),
            severity="critical",
        )
    elif target_status and (target_status.get("status") or "").strip().lower() != "ok":
        _append_unique_blocker(
            blockers,
            blocker_type="no_access",
            message=(
                target_status.get("error")
                or "Subscription access issues prevented ARM validation for this region."
            ),
            severity="critical",
        )
    elif not any_arm_access:
        _append_unique_blocker(
            blockers,
            blocker_type="no_access",
            message="No accessible subscriptions were available for ARM validation.",
            severity="critical",
        )

    sku_ready = True
    fallback_used = False
    restricted_notes: List[str] = []
    constrained_labels: List[str] = []
    for req in required_families:
        primary_family = str(req.get("primary_family") or "").strip()
        if not primary_family:
            continue
        primary_label = str(req.get("primary_label") or _short_label(primary_family)).strip()
        alt_family = str(req.get("alt_family") or "").strip() or None
        alt_label = str(req.get("alt_label") or _short_label(alt_family or "")).strip() or None

        primary = _merge_family_zone_health(
            region_short=region_short,
            family_ids=[primary_family],
            per_sub_results=per_sub_results,
        )
        alt = _merge_family_zone_health(
            region_short=region_short,
            family_ids=[alt_family] if alt_family else [],
            per_sub_results=per_sub_results,
        ) if alt_family else {"zones": None, "any_access": primary["any_access"], "found": False, "notes": []}

        for note in list(primary.get("notes") or []) + list(alt.get("notes") or []):
            _append_unique_message(restricted_notes, str(note))

        primary_zones = primary.get("zones")
        alt_zones = alt.get("zones")
        primary_all = bool(primary_zones) and all(primary_zones)
        alt_all = bool(alt_zones) and all(alt_zones)

        if primary_all:
            continue
        _append_unique_message(constrained_labels, primary_label)
        if alt_all and alt_label:
            fallback_used = True
            if primary_zones:
                _append_unique_blocker(
                    blockers,
                    blocker_type="zone_gap",
                    message=(
                        f"{primary_label} available in {_zones_label(primary_zones)}; "
                        f"using fallback {alt_label} across all 3 zones"
                    ),
                    severity="warning",
                )
            else:
                _append_unique_blocker(
                    blockers,
                    blocker_type="sku_unavailable",
                    message=(
                        f"{primary_label} unavailable in {region.get('display') or region.get('name') or region_short}; "
                        f"using fallback {alt_label}"
                    ),
                    severity="warning",
                )
            _append_unique_message(
                constraints,
                f"Using fallback SKU {alt_label} for {primary_label}",
            )
            continue

        sku_ready = False
        if not primary.get("any_access") and not alt.get("any_access"):
            _append_unique_blocker(
                blockers,
                blocker_type="no_access",
                message=f"Could not validate SKU availability for {primary_label}.",
                severity="critical",
            )
            continue
        if not primary_zones and not alt_zones:
            message = f"{primary_label} unavailable in this region"
            if alt_label:
                message += f"; fallback {alt_label} also unavailable"
            _append_unique_blocker(
                blockers,
                blocker_type="sku_unavailable",
                message=message,
                severity="critical",
            )
            continue
        message = f"{primary_label} available in {_zones_label(primary_zones)}"
        if alt_label:
            if alt_zones:
                message += f"; fallback {alt_label} available in {_zones_label(alt_zones)}"
            else:
                message += f"; fallback {alt_label} unavailable"
        message += " — no single family covers all 3 zones"
        _append_unique_blocker(
            blockers,
            blocker_type="zone_gap",
            message=message,
            severity="critical",
        )

    if sku_ready and not fallback_used:
        _append_unique_message(reasons, "All required SKU families available in all 3 zones")

    quota_tiers = region.get("quota_tiers") or {}
    family_quota = quota_tiers.get("families") or {}
    if not quota_required:
        _append_unique_message(reasons, "No quota requirement specified in BOM")
    elif region.get("quota_status") == "unknown":
        _append_unique_blocker(
            blockers,
            blocker_type="no_access",
            message="Quota status is unknown; validate vCPU headroom manually.",
            severity="critical",
        )
    else:
        any_quota_issues = False
        for family_info in family_quota.values():
            label = str(
                family_info.get("label")
                or family_info.get("family")
                or "Required family"
            ).strip()
            required_cores = int(family_info.get("required_cores") or 0)
            overall_status = family_info.get("overall_status")
            source = _quota_family_satisfaction_source(family_info)
            if overall_status == "insufficient":
                any_quota_issues = True
                best_headroom = None
                for candidate in [
                    family_info.get("tier1_sub_quota"),
                    family_info.get("tier2_quota_group"),
                ]:
                    headroom = _quota_num((candidate or {}).get("headroom"))
                    if headroom is not None and (best_headroom is None or headroom > best_headroom):
                        best_headroom = headroom
                suffix = ""
                if best_headroom is not None:
                    suffix = (
                        f" (need {required_cores} vCPU, best headroom "
                        f"{_format_quota_number(best_headroom)})"
                    )
                _append_unique_blocker(
                    blockers,
                    blocker_type="quota_insufficient",
                    message=f"Insufficient quota for {label}{suffix}",
                    severity="warning",
                )
            elif overall_status == "sufficient":
                tight_msg = _quota_tight_message(label, required_cores, source)
                if tight_msg:
                    _append_unique_message(constraints, tight_msg)
        if not any_quota_issues and region.get("quota_status") == "sufficient":
            _append_unique_message(reasons, "Quota sufficient for required SKU families")

    critical_non_validation = [
        b for b in blockers
        if b.get("severity") == "critical" and b.get("type") != "no_access"
    ]
    validation_blockers = [b for b in blockers if b.get("type") == "no_access"]
    warning_blockers = [b for b in blockers if b.get("severity") == "warning"]

    if critical_non_validation:
        verdict = "not_recommended"
    elif validation_blockers:
        verdict = "needs_validation"
    elif warning_blockers or constraints:
        verdict = "ready_with_constraints"
    else:
        verdict = "ready"

    if verdict == "ready":
        pass  # positive reasons already accumulated above
    elif verdict == "ready_with_constraints":
        pass  # constraints list speaks for itself
    # For negative verdicts, don't add summary reasons — blockers explain why

    region_display = str(
        region.get("display") or region.get("name") or region_short or "This region"
    ).strip()
    recommendations = _build_recommendations(
        blockers=blockers,
        region=region,
        region_display=region_display,
        fallback_used=fallback_used,
        restricted_notes=restricted_notes,
        constrained_labels=constrained_labels,
        quota_required=quota_required,
    )

    return {
        "verdict": verdict,
        "reasons": reasons,
        "blockers": blockers,
        "constraints": constraints,
        "recommendations": recommendations,
    }


def _merge_subscription_sku_records(
    *,
    subscriptions: List[Dict],
    per_sub_results: Dict[str, Dict],
    regions: Iterable[str],
    families: Iterable[str],
) -> List[Dict]:
    family_canonical = _canonical_family_map(families)
    want_regions = sorted({
        arm_sku_availability.arm_skus._normalize_region(r)  # noqa: SLF001
        for r in regions if r and str(r).strip()
    })
    merged: List[Dict] = []
    for region in want_regions:
        for fam_lower, fam_canonical in sorted(family_canonical.items()):
            accessible_rows: List[Dict] = []
            sub_details: Dict[str, Dict] = {}
            notes: List[str] = []
            for sub in subscriptions:
                sub_id = sub["subscription_id"]
                info = per_sub_results.get(sub_id) or {}
                row = (info.get("sku_index") or {}).get((region, fam_lower))
                detail = {
                    "status": info.get("status") or "unknown",
                    "error": info.get("error"),
                }
                if row:
                    detail.update({
                        "zones": list(row.get("zones") or [False, False, False]),
                        "sub_restricted": bool(row.get("sub_restricted")),
                        "sub_restriction_raw": row.get("sub_restriction_raw"),
                    })
                    if row.get("sub_restriction_raw"):
                        notes.append(str(row["sub_restriction_raw"]))
                sub_details[sub_id] = detail
                if info.get("status") == "ok" and row:
                    accessible_rows.append(row)

            if accessible_rows:
                zones = [
                    any(bool((row.get("zones") or [False, False, False])[i]) for row in accessible_rows)
                    for i in range(3)
                ]
                restricted_all = all(bool(row.get("sub_restricted")) for row in accessible_rows)
                merged_reason = "Available" if any(zones) else (
                    "Restricted in all checked subscriptions" if restricted_all else (notes[0] if notes else "Unavailable")
                )
                restricted_zones = [
                    str(i + 1)
                    for i in range(3)
                    if all(not bool((row.get("zones") or [False, False, False])[i]) for row in accessible_rows)
                ]
                merged.append({
                    "region": region,
                    "family": fam_canonical,
                    "display": arm_sku_availability._friendly_family(fam_canonical),  # noqa: SLF001
                    "zones": zones,
                    "sub_restricted": bool(restricted_all and not any(zones)),
                    "sub_restriction_raw": merged_reason,
                    "per_sub_results": sub_details,
                    "arm_provenance": {
                        "available": True,
                        "region_restricted": bool(restricted_all and not any(zones)),
                        "restricted_zones": restricted_zones,
                        "available_zones": [str(i + 1) for i, ok in enumerate(zones) if ok],
                        "notes": notes[:8],
                        "subscription_ids": [s["subscription_id"] for s in subscriptions],
                    },
                })
            else:
                merged.append({
                    "region": region,
                    "family": fam_canonical,
                    "display": arm_sku_availability._friendly_family(fam_canonical),  # noqa: SLF001
                    "zones": [False, False, False],
                    "sub_restricted": True,
                    "sub_restriction_raw": "No accessible subscriptions were available for ARM checks.",
                    "per_sub_results": sub_details,
                    "arm_provenance": {
                        "available": False,
                        "region_restricted": True,
                        "restricted_zones": ["1", "2", "3"],
                        "available_zones": [],
                        "notes": ["No accessible subscriptions were available for ARM checks."],
                        "subscription_ids": [s["subscription_id"] for s in subscriptions],
                    },
                })
    return merged


def _short_label(family_id: str) -> str:
    """standardDav6Family -> Dav6.  Best-effort default label."""
    import re as _re
    m = _re.match(r"^standard(.+)Family$", family_id, _re.IGNORECASE)
    return m.group(1) if m else family_id


# Limits for the user-facing modal override field. Kept here so handler and
# tests share the same constants.
MAX_FAMILIES_OVERRIDE_BYTES = 8 * 1024
MAX_FAMILIES_OVERRIDE_ROWS = 50


def parse_families_override(text: str) -> Optional[List[Dict]]:
    """Parse the multiline modal-override textarea into a list of
    required-family dicts. Returns None when the input is empty/blank so the
    caller falls back to the BOM sheet. Raises CompileError with a stable
    code on bad input."""
    if not text or not text.strip():
        return None
    if len(text) > MAX_FAMILIES_OVERRIDE_BYTES:
        raise CompileError(
            "bad_families_override",
            f"SKU override is too long ({len(text)} chars > {MAX_FAMILIES_OVERRIDE_BYTES}).",
            400,
        )
    out: List[Dict] = []
    seen = set()
    for lineno, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if len(out) >= MAX_FAMILIES_OVERRIDE_ROWS:
            raise CompileError(
                "bad_families_override",
                f"SKU override has too many rows (max {MAX_FAMILIES_OVERRIDE_ROWS}).",
                400,
            )
        if "\t" in line:
            parts = [p.strip() for p in line.split("\t")]
        else:
            parts = [p.strip() for p in line.split(",")]
        parts = (parts + ["", "", "", "", ""])[:5]
        pf, pl, af, al, cores_raw = parts
        if not pf:
            raise CompileError(
                "bad_families_override",
                f"SKU override line {lineno}: missing primary family ID.", 400)
        if pf.lower() in seen:
            raise CompileError(
                "bad_families_override",
                f"SKU override line {lineno}: duplicate primary family '{pf}'.", 400)
        seen.add(pf.lower())
        cores = _parse_required_cores(
            cores_raw, where=f"SKU override line {lineno}",
        )
        out.append({
            "primary_family": pf,
            "primary_label":  pl or None,
            "alt_family":     af or None,
            "alt_label":      al or None,
            "required_cores": cores,
        })
    if not out:
        return None
    return out


def _read_required_skus_sheet(wb) -> Optional[List[Dict]]:
    """Optional sheet on the BOM xlsx that lets each customer carry their own
    SKU family list with the BOM file. Returns None if the sheet is absent so
    callers can fall back to the static default. Raises CompileError if the
    sheet is present but malformed (better to fail loud than silently use
    defaults the user thought they'd overridden)."""
    if REQUIRED_SKUS_SHEET not in wb.sheetnames:
        return None
    ws = wb[REQUIRED_SKUS_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    # Find the header row by looking for "Primary Family" (case-insensitive).
    # This lets the sheet have an optional banner row above the header.
    header_idx = -1
    header: List[str] = []
    for i, raw in enumerate(rows):
        if raw is None:
            continue
        cells = [str(c or "").strip() for c in raw]
        if any(c.lower() == "primary family" for c in cells):
            header = [c.lower() for c in cells]
            header_idx = i
            break
    if header_idx < 0:
        raise CompileError(
            "bad_required_skus_sheet",
            f"'{REQUIRED_SKUS_SHEET}' sheet is present but has no 'Primary Family' header row.",
            400,
        )

    try:
        col_pf = header.index("primary family")
    except ValueError:
        raise CompileError(
            "bad_required_skus_sheet",
            f"'{REQUIRED_SKUS_SHEET}' sheet missing 'Primary Family' column.",
            400,
        )
    col_pl = header.index("primary label") if "primary label" in header else None
    col_af = header.index("alt family") if "alt family" in header else None
    col_al = header.index("alt label") if "alt label" in header else None
    col_rc = header.index("required cores") if "required cores" in header else None

    out: List[Dict] = []
    for raw in rows[header_idx + 1:]:
        if raw is None:
            # Blank gap row signals end-of-data; everything below is treated
            # as guidance/notes so authors can document the sheet inline.
            break
        cells = [("" if c is None else str(c)).strip() for c in raw]
        # All-blank row → end of the data block.
        if not any(cells):
            break
        if col_pf >= len(cells):
            continue
        pf = cells[col_pf]
        # Empty primary-family cell within the data block (uncommon — e.g. a
        # row with notes in column B but blank A) ends the block.
        if not pf:
            break
        # Defense-in-depth for users who delete the blank separator row when
        # editing the template: if the primary-family cell doesn't look
        # like a valid Azure family identifier (e.g. "Notes:", "•",
        # "- item"), treat it as the start of the notes section and stop.
        if not _FAMILY_ID_RE.match(pf):
            break
        pl = cells[col_pl] if (col_pl is not None and col_pl < len(cells)) else ""
        af = cells[col_af] if (col_af is not None and col_af < len(cells)) else ""
        al = cells[col_al] if (col_al is not None and col_al < len(cells)) else ""
        # Required Cores: read the RAW cell value (not the stringified copy)
        # so we keep numeric types when openpyxl returns them. Fall back to
        # the stringified copy if col_rc is out of range.
        cores_cell = None
        if col_rc is not None:
            if col_rc < len(raw):
                cores_cell = raw[col_rc]
            elif col_rc < len(cells):
                cores_cell = cells[col_rc]
        cores = _parse_required_cores(
            cores_cell, where=f"'{REQUIRED_SKUS_SHEET}' row {header_idx + 2 + len(out)} ({pf})",
        )
        out.append({
            "primary_family": pf,
            "primary_label": pl or _short_label(pf),
            "alt_family": af or None,
            "alt_label": (al or (_short_label(af) if af else None)),
            "required_cores": cores,
        })
    if not out:
        # Empty sheet (just headers) — treat as "use defaults" rather than
        # crashing the run.
        return None
    return out


def _read_bom_xlsx_bytes(blob: bytes) -> Tuple[List[str], List[Dict], Optional[List[Dict]]]:
    """Same logic as pipeline.sources.read_bom_v2 but from in-memory bytes.
    Returns (header, records, optional_required_families_from_sheet)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    except Exception as ex:
        raise CompileError("bad_step2_file",
                           f"Could not open Step 2 file as XLSX: {ex}", 400)
    if "Region Results" not in wb.sheetnames:
        raise CompileError("bad_step2_file",
                           "Step 2 file is missing the 'Region Results' sheet.", 400)
    ws = wb["Region Results"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise CompileError("bad_step2_file",
                           f"Step 2 file too small (only {len(rows)} rows).", 400)
    header = [c if c is not None else "" for c in rows[2]]
    if header[:3] != ["Region", "Display Name", "Overall Status"]:
        raise CompileError("bad_step2_file",
                           f"Step 2 unexpected header layout: {header[:3]}", 400)
    records = []
    for raw in rows[3:]:
        if not raw or not raw[0]:
            continue
        rec = {header[i]: raw[i] for i in range(len(header)) if header[i]}
        records.append(rec)

    required_skus = _read_required_skus_sheet(wb)
    return header, records, required_skus


def _load_latency() -> Dict:
    from . import dataset_store
    path = dataset_store.resolve_path("latency")
    return pipeline_sources.read_latency_csv(path)


def _flatten_family_ids(required_families: List[Dict]) -> List[str]:
    """Pull primary + alt family IDs out of a required_families list, dedup
    (case-insensitive), preserve order so ARM caching is stable."""
    out: List[str] = []
    seen = set()
    for entry in required_families:
        for k in ("primary_family", "alt_family"):
            v = (entry.get(k) or "").strip()
            if not v:
                continue
            kl = v.lower()
            if kl in seen:
                continue
            seen.add(kl)
            out.append(v)
    return out


def _validate_required_families(required: List[Dict]) -> None:
    """Sanity-check shape before we ship to ARM/model. Raises
    CompileError so the caller turns it into a 400."""
    if not isinstance(required, list) or not required:
        raise CompileError("bad_required_families",
                           "Required SKU families list is empty.", 400)
    seen = set()
    for i, entry in enumerate(required):
        if not isinstance(entry, dict):
            raise CompileError("bad_required_families",
                               f"Required family #{i+1} is not an object.", 400)
        pf = (entry.get("primary_family") or "").strip()
        if not pf:
            raise CompileError("bad_required_families",
                               f"Required family #{i+1} missing 'primary_family'.", 400)
        if not _FAMILY_ID_RE.match(pf):
            raise CompileError(
                "bad_required_families",
                (f"Required family #{i+1} has invalid primary_family "
                 f"{pf!r}. Expected an Azure family identifier like "
                 f"'standardDav6Family' (ASCII letters, digits, _, -)."),
                400,
            )
        if pf.lower() in seen:
            raise CompileError("bad_required_families",
                               f"Duplicate primary_family '{pf}' in required SKUs.", 400)
        seen.add(pf.lower())
        if not (entry.get("primary_label") or "").strip():
            entry["primary_label"] = _short_label(pf)
        af = (entry.get("alt_family") or "").strip()
        if af and not _FAMILY_ID_RE.match(af):
            raise CompileError(
                "bad_required_families",
                (f"Required family #{i+1} has invalid alt_family "
                 f"{af!r}. Expected an Azure family identifier like "
                 f"'standardDASv5Family' (ASCII letters, digits, _, -)."),
                400,
            )
        entry["alt_family"] = af or None
        if af and not (entry.get("alt_label") or "").strip():
            entry["alt_label"] = _short_label(af)
        # Primary/alt labels MUST differ when an alt is present — the
        # recommendation engine uses label-equality to distinguish which one
        # was chosen, so identical labels would misclassify fallbacks as
        # primary and silently corrupt the dashboard's filter/overview math.
        if af and entry.get("primary_label") == entry.get("alt_label"):
            raise CompileError(
                "bad_required_families",
                (f"Required family #{i+1} has identical primary_label and "
                 f"alt_label ({entry.get('primary_label')!r}). When an "
                 f"alt_family is set, the two labels must differ so the "
                 f"dashboard can tell which one a region picked."),
                400,
            )
        # Normalize required_cores: any incoming type goes through
        # _parse_required_cores so the downstream contract (None | positive
        # int) holds regardless of input source (sheet, override, JSON API).
        entry["required_cores"] = _parse_required_cores(
            entry.get("required_cores"),
            where=f"Required family #{i+1} ({pf})",
        )


def _global_unscoped_note() -> str:
    return (
        "Per-subscription restrictions not evaluated — operator lacks Reader on target subscription"
    )


def compile_snapshot(
    *,
    subscription_id: Optional[str] = None,
    arm_token: Optional[str] = None,
    subscriptions: Optional[List[Dict]] = None,
    step2_bytes: Optional[bytes] = None,
    bom_data: Optional[Dict] = None,
    customer_segments: Optional[List[str]] = None,
    customer_name: Optional[str] = None,
    regions: Optional[List[str]] = None,
    families: Optional[List[str]] = None,
    required_families_override: Optional[List[Dict]] = None,
    triggered_by_email: str,
    triggered_by_oid: str,
    source_label: str = "upload+live-arm",
    run_id: Optional[str] = None,
    progress_token: Optional[str] = None,
) -> Dict:
    """
    Run a full compile and return the snapshot dict (not yet persisted).

    Exactly one of ``step2_bytes`` (legacy xlsx upload) or ``bom_data``
    (in-app BOM editor) must be supplied. ``bom_data`` is a dict with:
      * ``bom_header``: list[str] — first three cols are
        ["Region","Display Name","Overall Status"]; remaining cols are
        service names.
      * ``bom_records``: list[dict] — one row per region, keyed by header,
        values are the ✅/❌ marker strings.
      * ``required_families``: list[dict] — already-validated family
        records (primary_family, primary_label, alt_family, alt_label,
        required_cores). When provided, this short-circuits the
        BOM/sheet/default resolution priority.

    Required-SKU resolution priority (first match wins):
      1. ``required_families_override`` arg (explicit override path)
      2. ``bom_data["required_families"]`` when ``bom_data`` is supplied
      3. The "Required SKUs" sheet inside ``step2_bytes``
      4. Default families from ``api/_shared/data/skus.txt`` paired with
         the canonical labels in ``pipeline_model.DEFAULT_REQUIRED_FAMILIES``.

    Raises CompileError with stable codes the caller surfaces to the UI.
    """
    t_start = time.time()

    # Progress: phase 0 = Loading BOM. Set first so the modal sees a
    # transition out of percent=0 quickly even though this phase is fast.
    if progress_token:
        run_progress.set_phase(progress_token, 0, label="Loading BOM")

    subscriptions = _normalize_subscriptions(
        subscription_id=subscription_id,
        arm_token=arm_token,
        subscriptions=subscriptions,
    )
    subscription_id = subscriptions[0]["subscription_id"]
    if step2_bytes and bom_data:
        raise CompileError(
            "bom_source_conflict",
            ("Provide either step2_bytes (xlsx upload) or bom_data "
             "(in-app BOM), not both."),
            400,
        )
    if not step2_bytes and not bom_data:
        raise CompileError(
            "missing_step2",
            "A Step 2 BOM file (region_results_*.xlsx) or in-app BOM is required.",
            400,
        )

    regions = regions or arm_sku_availability.load_default_regions(DATA_DIR)
    customer_name = (customer_name or "").strip() or None

    if bom_data is not None:
        bom_header = bom_data.get("bom_header") or [
            "Region", "Display Name", "Overall Status",
        ]
        bom_records = bom_data.get("bom_records") or []
        bom_required = bom_data.get("required_families") or None
        bom_services_selected = bom_data.get("services") or []
        log.info("in-app BOM ok: %d records (required-skus from bom_data: %s)",
                 len(bom_records), "yes" if bom_required else "no")
    else:
        bom_header, bom_records, bom_required = _read_bom_xlsx_bytes(step2_bytes)
        bom_services_selected = []
        log.info("step2 ok: %d records (required-skus sheet: %s)",
                 len(bom_records), "yes" if bom_required else "no")

    skus_source: str
    if required_families_override:
        required_families = required_families_override
        skus_source = "modal_override"
    elif bom_required:
        required_families = bom_required
        skus_source = "bom_data" if bom_data is not None else "bom_sheet"
    else:
        # Pair the static skus.txt order with the canonical labels in
        # DEFAULT_REQUIRED_FAMILIES so the v6/v5 fallback narrative still
        # works for the legacy default set.
        required_families = list(pipeline_model.DEFAULT_REQUIRED_FAMILIES)
        skus_source = "skus_txt_default"
    _validate_required_families(required_families)

    # Build the flat family-id list for ARM. If the caller
    # explicitly passed `families` (back-compat / tests), honor that too.
    families = families or _flatten_family_ids(required_families)
    if not families:
        raise CompileError("bad_required_families",
                           "Resolved 0 SKU families to query ARM with.", 400)

    log.info("compile start sub=%s regions=%d families=%d subs=%d (source=%s)",
             subscription_id, len(regions), len(families), len(subscriptions), skus_source)

    t_arm = time.time()
    if progress_token:
        run_progress.set_phase(
            progress_token, 1, label="ARM SKU availability",
        )
    per_sub_results: Dict[str, Dict] = {}
    per_sub_status: Dict[str, Dict] = {}
    for sub in subscriptions:
        sub_id = sub["subscription_id"]
        sub_token = sub.get("arm_token")
        sub_status = sub.get("status") or ("ok" if sub_token else "no_access")
        sub_role = sub.get("role") or "auxiliary"
        t_sub = time.time()
        activity_log.record(
            "arm_call_start",
            actor_email=triggered_by_email, actor_oid=triggered_by_oid,
            subscription_id=sub_id, run_id=run_id,
            api_scope="subscription",
            message=(f"ARM Microsoft.Compute/skus — subscription {sub_id} "
                     f"× {len(regions)} regions × {len(families)} families"),
            details={
                "subscription_id": sub_id,
                "provider": "Microsoft.Compute",
                "operation": "resourceSkus/list",
                "regions": regions,
                "families": families,
                "token_fingerprint": activity_log.token_fingerprint(sub_token),
            },
        )
        if sub_status != "ok" or not sub_token:
            reason = str(sub.get("error") or "No ARM access for this subscription.")
            sku_rows = _placeholder_sku_records(
                subscription_id=sub_id,
                regions=regions,
                families=families,
                reason=reason,
                status="no_access",
            )
            quota_result = {
                "subscription_id": sub_id,
                "status": "no_access",
                "error": reason,
                "has_quota_groups": False,
                "groups": [],
            }
            sub_quota = {
                "subscription_id": sub_id,
                "status": "no_access",
                "error": reason,
                "regions": {},
            }
            per_sub_status[sub_id] = {"status": "no_access", "error": reason}
            per_sub_results[sub_id] = {
                "status": "no_access",
                "error": reason,
                "sku_records": sku_rows,
                "sku_index": _index_sku_records(sku_rows),
                "quota_groups": quota_result,
                "subscription_quota": sub_quota,
            }
            activity_log.record(
                "arm_call_error",
                actor_email=triggered_by_email, actor_oid=triggered_by_oid,
                subscription_id=sub_id, run_id=run_id,
                api_scope="subscription", status="error",
                message=reason,
                duration_ms=int((time.time() - t_sub) * 1000),
            )
            continue
        try:
            sku_rows = arm_sku_availability.fetch_arm_sku_records(
                arm_token=sub_token,
                subscription_id=sub_id,
                want_regions=regions,
                want_families=families,
            )
            quota_result = quota_groups.check_quota_groups(
                sub_token, sub_id, regions, families,
            )
            sub_quota = quota_groups.check_subscription_quota(
                sub_token, sub_id, regions, families,
            )
            per_sub_status[sub_id] = {"status": "ok"}
            per_sub_results[sub_id] = {
                "status": "ok",
                "error": None,
                "sku_records": sku_rows,
                "sku_index": _index_sku_records(sku_rows),
                "quota_groups": quota_result,
                "subscription_quota": sub_quota,
            }
            activity_log.record(
                "arm_call_ok",
                actor_email=triggered_by_email, actor_oid=triggered_by_oid,
                subscription_id=sub_id, run_id=run_id,
                api_scope="subscription", status="ok",
                message=f"ARM returned {len(sku_rows)} SKU availability rows",
                duration_ms=int((time.time() - t_sub) * 1000),
                details={"rows_returned": len(sku_rows)},
            )
        except arm_skus.ArmError as ex:
            if ex.status == 401 and sub_role == "target":
                activity_log.record(
                    "arm_call_error",
                    actor_email=triggered_by_email, actor_oid=triggered_by_oid,
                    subscription_id=sub_id, run_id=run_id,
                    api_scope="subscription", status="error",
                    message=f"{ex.code}: {ex.message}",
                    duration_ms=int((time.time() - t_sub) * 1000),
                )
                raise CompileError(f"arm_{ex.code}", ex.message, ex.status)
            status = "no_access" if ex.status in (401, 403) else "error"
            # Provide a concise reason for the placeholder records — the raw
            # ArmError message includes per-region details (+55 more…) that
            # clutter the UI when shown in every cell.
            if ex.status == 403:
                reason = f"No read access on subscription {sub_id}. Assign Reader role to query SKU availability."
            elif ex.status == 401:
                reason = f"Token expired or invalid for subscription {sub_id}."
            else:
                reason = ex.message
            sku_rows = _placeholder_sku_records(
                subscription_id=sub_id,
                regions=regions,
                families=families,
                reason=reason,
                status=status,
            )
            quota_result = {
                "subscription_id": sub_id,
                "status": status,
                "error": reason,
                "has_quota_groups": False,
                "groups": [],
            }
            sub_quota = {
                "subscription_id": sub_id,
                "status": status,
                "error": reason,
                "regions": {},
            }
            per_sub_status[sub_id] = {"status": status, "error": reason}
            per_sub_results[sub_id] = {
                "status": status,
                "error": reason,
                "sku_records": sku_rows,
                "sku_index": _index_sku_records(sku_rows),
                "quota_groups": quota_result,
                "subscription_quota": sub_quota,
            }
            activity_log.record(
                "arm_call_error",
                actor_email=triggered_by_email, actor_oid=triggered_by_oid,
                subscription_id=sub_id, run_id=run_id,
                api_scope="subscription", status="error",
                message=f"{ex.code}: {ex.message}",
                duration_ms=int((time.time() - t_sub) * 1000),
            )

    if len(subscriptions) == 1:
        only_sub = subscriptions[0]["subscription_id"]
        sku_records = list((per_sub_results.get(only_sub) or {}).get("sku_records") or [])
    else:
        sku_records = _merge_subscription_sku_records(
            subscriptions=subscriptions,
            per_sub_results=per_sub_results,
            regions=regions,
            families=families,
        )
    target_status = per_sub_status.get(subscription_id) or {}
    accessible_helpers = [
        s["subscription_id"]
        for s in subscriptions
        if s.get("role") != "target"
        and (per_sub_status.get(s["subscription_id"]) or {}).get("status") == "ok"
    ]
    mode = "subscription_scoped"
    mode_note = None
    mode_reason = None
    sku_query_subscription_id = subscription_id
    if accessible_helpers and target_status.get("status") != "ok":
        mode = "global_unscoped"
        mode_note = _global_unscoped_note()
        mode_reason = target_status.get("error")
        sku_query_subscription_id = accessible_helpers[0]
        activity_log.record(
            "arm_call_skipped",
            actor_email=triggered_by_email, actor_oid=triggered_by_oid,
            subscription_id=subscription_id, run_id=run_id,
            api_scope="subscription", status="warn",
            message=(f"Using global ARM SKU mode via operator subscription "
                     f"{sku_query_subscription_id}"),
            details={
                "mode": mode,
                "note": mode_note,
                "reason": mode_reason,
                "target_subscription_id": subscription_id,
                "query_subscription_id": sku_query_subscription_id,
            },
        )
    log.info("arm sku availability ok: %d merged records across %d subscriptions",
             len(sku_records), len(subscriptions))

    latency = _load_latency()

    if progress_token:
        run_progress.set_phase(
            progress_token, 2, label="Building model",
        )

    raw = {
        "sku_records": sku_records,
        "bom_header": bom_header,
        "bom_records": bom_records,
        "latency": latency,
    }
    try:
        m = pipeline_model.build_model(raw, data_dir=None,
                                       required_families=required_families)
    except RuntimeError as ex:
        raise CompileError("model_error", str(ex), 422)

    # Persist the build_model inputs so the dashboard's BOM & SKUs tab can
    # render them. Without this they only exist at compile time and the tab
    # would always render an "unavailable in this snapshot" empty state.
    m["bom_header"] = bom_header
    m["bom_records"] = bom_records
    m["sku_records"] = sku_records
    m["per_sub_results"] = {
        sub_id: {
            "status": info.get("status"),
            "error": info.get("error"),
            "sku_records": info.get("sku_records") or [],
            "quota_groups": info.get("quota_groups"),
            "subscription_quota": info.get("subscription_quota"),
        }
        for sub_id, info in per_sub_results.items()
    }
    m["quota_groups"] = {
        sub_id: info.get("quota_groups")
        for sub_id, info in per_sub_results.items()
    }
    tiered_quota = _evaluate_quota_tiers(
        per_sub_results=per_sub_results,
        required_families=required_families,
        regions=regions,
        target_subscription_id=subscription_id,
    )
    _apply_quota_group_status(
        snapshot_regions=m.get("regions") or [],
        required_families=required_families,
        per_sub_results=per_sub_results,
        per_sub_status=per_sub_status,
        tiered_results=tiered_quota,
    )
    m["quota_remediation"] = _build_quota_remediation_plan(
        per_sub_results=per_sub_results,
        quota_tiers=tiered_quota,
        regions=m.get("regions") or [],
        required_families=required_families,
    )

    snapshot_meta = {
        "subscription_id": subscription_id,
        "subscription_ids": [s["subscription_id"] for s in subscriptions],
        "per_sub_status": per_sub_status,
        "mode": mode,
        "mode_note": mode_note,
        "mode_reason": mode_reason,
        "target_subscription_id": subscription_id,
        "sku_query_subscription_id": sku_query_subscription_id,
        "customer_name": customer_name,
        "customer_segments": customer_segments,
        "triggered_by_email": triggered_by_email,
        "triggered_by_oid": triggered_by_oid,
        "source": source_label,
        "engine_version": ENGINE_VERSION,
        "regions_requested": len(regions),
        "families_requested": len(families),
        "skus_source": skus_source,
        "skus_resolved": required_families,
        "services": bom_services_selected,
        "compile_seconds": round(time.time() - t_start, 2),
        "sku_availability": {
            "provider": "arm",
            "subscription_id": subscription_id,
            "subscription_ids": [s["subscription_id"] for s in subscriptions],
            "target_subscription_id": subscription_id,
            "query_subscription_id": sku_query_subscription_id,
            "regions_queried": len(regions),
            "rows_returned": len(sku_records),
            "applied": True,
            "mode": mode,
            "note": mode_note,
            "reason": mode_reason,
            "per_subscription_restrictions_evaluated": mode != "global_unscoped",
        },
    }
    m["meta"] = snapshot_meta
    for region in m.get("regions") or []:
        region["deployment_verdict"] = _compute_deployment_verdict(
            region,
            snapshot_meta=snapshot_meta,
            per_sub_results=per_sub_results,
        )
    return m


# ---- Persistence -----------------------------------------------------------

def persist_snapshot(snapshot: Dict, *, run_id: str, bom_id: Optional[str] = None) -> str:
    """Write the snapshot blob and return the blob name. Snapshots are
    partitioned by ``bom_id`` so each BOM owns its own run history. Legacy
    callers that omit ``bom_id`` fall back to the subscription id, matching
    the pre-decoupling on-disk layout."""
    sub_id = snapshot["meta"]["subscription_id"]
    partition = (bom_id or sub_id)
    blob_name = f"{partition}/{run_id}.json"
    container = storage.get_blob_container("snapshots")
    payload = json.dumps(snapshot, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    container.upload_blob(name=blob_name, data=payload, overwrite=False,
                          content_type="application/json")
    return blob_name


def insert_run(run_id: str, *, status: str, subscription_id: str,
               triggered_by_email: str, triggered_by_oid: str,
               bom_id: Optional[str] = None,
               source: str = "upload+live-arm",
               error: Optional[str] = None,
               snapshot_blob: Optional[str] = None,
               customer_name: Optional[str] = None,
               customer_segments: Optional[List[str]] = None,
               arm_overlay: Optional[bool] = None) -> None:
    """Insert/update a row in the `runs` Table. Rows are partitioned by
    ``bom_id`` (the owning BOM); ``subscription_id`` is stored as a column.
    Legacy callers that omit ``bom_id`` partition by subscription, matching
    the pre-decoupling layout."""
    table = storage.get_table_client("runs")
    now = datetime.now(timezone.utc).isoformat()
    entity = {
        "PartitionKey": (bom_id or subscription_id),
        "RowKey": run_id,
        "subscription_id": subscription_id,
        "status": status,
        "source": source,
        "triggered_by_email": triggered_by_email,
        "triggered_by_oid": triggered_by_oid,
        "started_at": now,
    }
    if status in ("succeeded", "failed"):
        entity["ended_at"] = now
    if error is not None:
        entity["error"] = error[:1024]
    if snapshot_blob is not None:
        entity["snapshot_blob"] = snapshot_blob
    if customer_name is not None:
        entity["customer_name"] = customer_name
    if customer_segments is not None:
        entity["customer_segments"] = ",".join(customer_segments)
    if arm_overlay is not None:
        entity["arm_overlay_applied"] = bool(arm_overlay)
    try:
        table.upsert_entity(entity, mode="merge")
    except Exception:
        log.exception("failed to upsert run %s", run_id)


def new_run_id() -> str:
    """`<UTC-ts>-<short-uuid>` — sortable AND unique."""
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
    return f"{ts}-{uuid.uuid4().hex[:8]}"
