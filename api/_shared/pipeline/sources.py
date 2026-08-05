"""
Source-file readers.

The optional file-based import flow reads three drop-in folders under a
local ``Development\\`` tree:

  Development\\Step 1 - SKU - Zone Dump\\sku-availability-<id>-<ts>.xlsx
  Development\\Step 2 - BOM Dump\\region_results_<ts>.xlsx
  Development\\Step 3 - Azure Region Latency Dump\\azure_region_latency.csv

When multiple files exist in a Step folder we pick the one with the most
recent timestamp encoded in the filename (NOT mtime, which is unreliable
under file-sync clients).

All source files are clean OOXML / CSV - no Excel COM required.
"""
from __future__ import annotations

import csv
import os
import re
from typing import Dict, List, Optional, Tuple

import openpyxl


# ----- Filename timestamp parsing --------------------------------------------

# sku-availability-<subId>-YYYYMMDD-HHMMSS.xlsx
_SKU_TS = re.compile(r"-(\d{8})-(\d{6})\.xlsx$", re.IGNORECASE)
# region_results_YYYYMMDD_HHMMSS.xlsx
_BOM_TS = re.compile(r"_(\d{8})_(\d{6})\.xlsx$", re.IGNORECASE)


def _parse_ts(name: str, pattern: re.Pattern) -> Optional[str]:
    m = pattern.search(name)
    if not m:
        return None
    return m.group(1) + m.group(2)


def _newest_by_filename_ts(directory: str, pattern: re.Pattern) -> Optional[str]:
    """Return the path to the file with the latest YYYYMMDDHHMMSS in its name."""
    if not os.path.isdir(directory):
        return None
    best: Tuple[Optional[str], Optional[str]] = (None, None)
    for entry in os.listdir(directory):
        full = os.path.join(directory, entry)
        if not os.path.isfile(full):
            continue
        ts = _parse_ts(entry, pattern)
        if ts is None:
            continue
        if best[0] is None or ts > best[0]:
            best = (ts, full)
    return best[1]


# ----- SKU file reader -------------------------------------------------------

def read_sku_v2(path: str) -> List[List]:
    """
    Read the new SKU availability dump.

    Returns a list of dict rows with keys:
        region, family, display, zones[3], sub_restricted (bool), sub_restriction_raw
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["SKU Availability"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    required = {"Region", "Family", "Display Name",
                "Zone 1", "Zone 2", "Zone 3",
                "Subscription Restriction"}
    missing = required - set(col)
    if missing:
        raise RuntimeError(f"SKU file {path} is missing columns: {sorted(missing)}")

    out: List[Dict] = []
    for raw in rows[1:]:
        if not raw or raw[col["Region"]] is None:
            continue
        region = str(raw[col["Region"]]).strip().lower()
        family = str(raw[col["Family"]] or "").strip()
        if not region or not family:
            continue

        sub_restriction_raw = str(raw[col["Subscription Restriction"]] or "").strip()
        sub_restricted = sub_restriction_raw.lower() != "available"

        # Zone N already encodes per-zone availability for THIS subscription;
        # the Subscription Restriction column is just the human-readable reason.
        # Don't AND with sub_restricted - that would over-zero zones for cases
        # like "Restricted in Zone 1" where zones 2/3 are actually Yes.
        z = [str(raw[col[k]] or "").strip().lower() == "yes"
             for k in ("Zone 1", "Zone 2", "Zone 3")]

        out.append({
            "region": region,
            "family": family,
            "display": str(raw[col["Display Name"]] or "").strip(),
            "zones": z,
            "sub_restricted": sub_restricted,
            "sub_restriction_raw": sub_restriction_raw,
        })
    return out


# ----- BOM (Region Results v2) reader ---------------------------------------

def read_bom_v2(path: str) -> Tuple[List[str], List[Dict]]:
    """
    Read the new region_results workbook. Returns (header, records).
    Layout: row1=banner, row2=legend, row3=headers, row4+=data.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Region Results"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise RuntimeError(f"BOM file {path} too small (only {len(rows)} rows)")

    header = [c if c is not None else "" for c in rows[2]]
    if header[:3] != ["Region", "Display Name", "Overall Status"]:
        raise RuntimeError(
            f"BOM file {path} unexpected header layout: {header[:3]}"
        )

    records: List[Dict] = []
    for raw in rows[3:]:
        if not raw or not raw[0]:
            continue
        rec = {header[i]: raw[i] for i in range(len(header)) if header[i]}
        records.append(rec)
    return header, records


# ----- Latency reader (unchanged) -------------------------------------------

def read_latency_csv(path: str) -> Dict[str, Dict[str, int]]:
    out: Dict[str, Dict[str, int]] = {}
    if not os.path.exists(path):
        return out
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader)
        dest_names = [h.strip() for h in headers[1:]]
        for row in reader:
            if not row or not row[0]:
                continue
            src = row[0].strip()
            for i, val in enumerate(row[1:]):
                v = (val or "").strip()
                if not v:
                    continue
                try:
                    ms = int(v)
                except ValueError:
                    continue
                out.setdefault(src, {})[dest_names[i]] = ms
    return out


# ----- Top-level loader ------------------------------------------------------

def _resolve_paths(project_root: str) -> Dict[str, str]:
    """Locate the three latest source files within the Development tree."""
    dev = os.path.join(project_root, "Development")

    sku_dir = os.path.join(dev, "Step 1 - SKU - Zone Dump")
    bom_dir = os.path.join(dev, "Step 2 - BOM Dump")
    lat_dir = os.path.join(dev, "Step 3 - Azure Region Latency Dump")

    sku_path = _newest_by_filename_ts(sku_dir, _SKU_TS)
    bom_path = _newest_by_filename_ts(bom_dir, _BOM_TS)

    if sku_path is None:
        raise FileNotFoundError(
            f"No SKU file matching sku-availability-*-YYYYMMDD-HHMMSS.xlsx "
            f"found in {sku_dir}"
        )
    if bom_path is None:
        raise FileNotFoundError(
            f"No BOM file matching region_results_YYYYMMDD_HHMMSS.xlsx "
            f"found in {bom_dir}"
        )

    lat_path = os.path.join(lat_dir, "azure_region_latency.csv")
    if not os.path.exists(lat_path):
        raise FileNotFoundError(f"Latency CSV not found at {lat_path}")

    return {"sku": sku_path, "bom": bom_path, "latency": lat_path}


def load_all(project_root: str) -> Dict:
    """
    Locate and read every source file. Returns dict consumed by model.build_model.
    """
    paths = _resolve_paths(project_root)
    sku_records = read_sku_v2(paths["sku"])
    bom_header, bom_records = read_bom_v2(paths["bom"])
    latency = read_latency_csv(paths["latency"])

    return {
        "paths": paths,
        "sku_records": sku_records,
        "bom_header": bom_header,
        "bom_records": bom_records,
        "latency": latency,
    }
