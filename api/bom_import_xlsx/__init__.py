"""POST /api/bom/import_xlsx

Server-side parse of either a `bom_template.xlsx` (BOM + Catalog +
Required SKUs sheets) or a `region_results_*.xlsx` (Region Results +
Required SKUs sheets). Returns JSON the frontend can hand directly to
the BOM editor for review + save — this endpoint does NOT persist
anything; the user must hit PUT /subscription_metadata/{sub_id} after.

multipart/form-data:
    file              : .xlsx (required)

Response:
    {
      "customer_name": "Avaya" | null,
      "services": [{"name": "Azure Automation"}, ...],
      "required_skus": [{primary_family, primary_label, alt_family, alt_label, required_cores}, ...],
      "source_format": "bom_template" | "region_results",
      "warnings": ["..."]
    }
"""
from __future__ import annotations

import io
import json
import logging
import re
from typing import Dict, List, Optional, Tuple

from .._shared import httpfunc as func
import openpyxl

from .._shared import activity_log, auth, bom_services, compile as compile_mod, csrf

log = logging.getLogger(__name__)

MAX_FILE_BYTES = 10 * 1024 * 1024


def _err(code: str, message: str, status: int = 400) -> func.HttpResponse:
    return func.HttpResponse(
        json.dumps({"error": code, "message": message}),
        status_code=status, mimetype="application/json",
    )


def _read_bom_sheet_services(wb, catalog_by_name: Dict[str, Dict]) -> Tuple[List[Dict], List[str]]:
    """Reads the 'BOM' sheet (column A, from row 4) used by the original
    bom_template.xlsx layout. Returns (services_list, warnings).

    Services are filtered against the catalog so unknown names surface
    as warnings rather than silently being saved.
    """
    warnings: List[str] = []
    if "BOM" not in wb.sheetnames:
        return [], warnings
    ws = wb["BOM"]
    services: List[Dict] = []
    seen: set = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[0] if row and len(row) > 0 else None
        if not name:
            continue
        n = str(name).strip()
        if not n:
            continue
        if n in seen:
            continue
        seen.add(n)
        if n in catalog_by_name:
            services.append({"name": n})
        else:
            warnings.append(
                f"Service '{n}' from the BOM sheet isn't in the in-app "
                f"service catalog — skipped."
            )
    return services, warnings


def _read_region_results_services(wb, catalog_by_name: Dict[str, Dict]) -> Tuple[List[Dict], List[str]]:
    """Reads the per-service columns from the 'Region Results' sheet header
    (columns 4+). This is what check_azure_regions.py produces.
    """
    warnings: List[str] = []
    if "Region Results" not in wb.sheetnames:
        return [], warnings
    ws = wb["Region Results"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return [], warnings
    header = rows[2]  # row index 2 = the header row
    services: List[Dict] = []
    seen: set = set()
    for col in header[3:]:
        if not col:
            continue
        n = str(col).strip()
        if not n or n in seen:
            continue
        seen.add(n)
        if n in catalog_by_name:
            services.append({"name": n})
        else:
            warnings.append(
                f"Service '{n}' from the Region Results header isn't in "
                f"the in-app service catalog — skipped."
            )
    return services, warnings


def _detect_customer_name(filename: Optional[str]) -> Optional[str]:
    """Pull customer name from filename like 'region_results_avaya.xlsx'."""
    if not filename:
        return None
    m = re.search(r"region_results_([A-Za-z0-9_\-]+)\.xlsx$", filename, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def main(req: func.HttpRequest) -> func.HttpResponse:
    principal = auth.get_local_user(req)

    def _log(status: str, message: str, **details) -> None:
        activity_log.record(
            "bom_import",
            actor_email=principal.email,
            actor_oid=principal.oid,
            api_scope="local",
            status=status,
            message=message,
            details=details or None,
        )

    try:
        csrf.assert_safe_origin(req)
    except csrf.OriginError as ex:
        log.warning("origin check rejected bom_import_xlsx: %s", ex)
        _log("error", "BOM import rejected by origin policy", code="origin_rejected")
        return _err("origin_rejected", str(ex), 403)

    try:
        files = req.files
    except Exception as ex:
        _log("error", "BOM import multipart parse failed", code="bad_request")
        return _err("bad_request", f"Could not parse multipart body: {ex}", 400)
    f = files.get("file")
    if f is None:
        _log("error", "BOM import missing file upload", code="missing_file")
        return _err("missing_file", "Upload an xlsx file in the 'file' field.", 400)
    blob = f.read()
    if not blob:
        _log("error", "BOM import uploaded file was empty", filename=getattr(f, "filename", ""))
        return _err("missing_file", "Uploaded file is empty.", 400)
    if len(blob) > MAX_FILE_BYTES:
        _log(
            "error",
            "BOM import file exceeded size limit",
            filename=getattr(f, "filename", ""),
            bytes=len(blob),
        )
        return _err(
            "file_too_large",
            f"File exceeds {MAX_FILE_BYTES // (1024 * 1024)} MB.",
            413,
        )
    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    except Exception as ex:
        _log("error", "BOM import workbook open failed", filename=getattr(f, "filename", ""))
        return _err("bad_xlsx", f"Could not open file as XLSX: {ex}", 400)

    catalog_by_name = bom_services.catalog_by_name()
    warnings: List[str] = []
    services: List[Dict] = []
    source_format = "unknown"

    if "BOM" in wb.sheetnames and "Catalog" in wb.sheetnames:
        services, w = _read_bom_sheet_services(wb, catalog_by_name)
        warnings.extend(w)
        source_format = "bom_template"
    elif "Region Results" in wb.sheetnames:
        services, w = _read_region_results_services(wb, catalog_by_name)
        warnings.extend(w)
        source_format = "region_results"
    else:
        _log("error", "BOM import workbook format not recognized", filename=getattr(f, "filename", ""))
        return _err(
            "bad_xlsx",
            ("Unrecognized workbook. Expected either 'BOM' + 'Catalog' "
             "(bom_template.xlsx) or 'Region Results' (region_results_*.xlsx)."),
            400,
        )

    required_skus: List[Dict] = []
    try:
        sheet_rows = compile_mod._read_required_skus_sheet(wb)
        if sheet_rows:
            required_skus = sheet_rows
    except compile_mod.CompileError as ex:
        warnings.append(f"'Required SKUs' sheet skipped: {ex.message}")

    customer_name = _detect_customer_name(getattr(f, "filename", None))

    payload = {
        "customer_name": customer_name,
        "services": services,
        "required_skus": required_skus,
        "source_format": source_format,
        "warnings": warnings,
    }
    _log(
        "ok",
        f"Imported BOM workbook ({source_format})",
        filename=getattr(f, "filename", ""),
        source_format=source_format,
        customer_name=customer_name,
        service_count=len(services),
        required_sku_count=len(required_skus),
        warning_count=len(warnings),
    )
    return func.HttpResponse(
        json.dumps(payload, ensure_ascii=False),
        status_code=200, mimetype="application/json",
    )
